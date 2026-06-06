"""Tests for the report-archetype layer (finance report shape upgrade)."""

from openpyxl import load_workbook

from pbc_chaos.config_loader import config_from_mapping, load_config
from pbc_chaos.generators.ap_aging import APAgingGenerator
from pbc_chaos.generators.ar_aging import ARAgingGenerator
from pbc_chaos.generators.base import CompanyProfile, FinancialPeriod
from pbc_chaos.pbc_workbook import _generate_documents, generate_pbc_workbook_with_ground_truth
from pbc_chaos.scoring import compare_extraction
from pbc_chaos.workbook import formatting
from pbc_chaos.workbook.report_archetypes import (
    SOFTWARE_SIGNATURES,
    build_report_frame,
    choose_software_signature,
)


def company_and_period():
    return (
        CompanyProfile(company_id="client_001", company_name="ABC Sdn Bhd"),
        FinancialPeriod.calendar_year(2025),
    )


def _ap_document(seed=7):
    company, period = company_and_period()
    return APAgingGenerator().generate(company=company, period=period, seed=seed)


def test_ap_summary_pivot_collapses_to_one_row_per_vendor():
    company, period = company_and_period()
    document = _ap_document()

    frame = build_report_frame(
        document=document,
        company=company,
        period=period,
        sheet_name="AP Aging",
        seed=1,
        severity=5,
    )

    assert frame.report_form == "summary_pivot"
    assert frame.grouping_key == "vendor_id"
    distinct_vendors = document.data["vendor_id"].nunique()
    assert len(frame.aggregated_canonical_view) == distinct_vendors
    # Canonical open items are never mutated.
    assert tuple(document.data.columns)[:5] == (
        "client_id",
        "financial_year",
        "period_start",
        "period_end",
        "currency",
    )


def test_ap_summary_pivot_grand_total_ties_out():
    company, period = company_and_period()
    document = _ap_document()

    frame = build_report_frame(
        document=document,
        company=company,
        period=period,
        sheet_name="AP Aging",
        seed=1,
        severity=5,
    )

    aggregated = frame.aggregated_canonical_view
    table = frame.table
    total_row = table.iloc[-1]
    assert total_row["vendor_name"] == "Grand Total"

    bucket_fields = [
        "current_amount",
        "bucket_1_30",
        "bucket_31_60",
        "bucket_61_90",
        "bucket_over_90",
    ]
    # Grand total foots every bucket and the balance.
    for field in bucket_fields + ["outstanding_amount"]:
        assert round(float(total_row[field]), 2) == round(float(aggregated[field].sum()), 2)

    # Each vendor balance equals the sum of its buckets (buckets partition the balance).
    bucket_sum = aggregated[bucket_fields].sum(axis=1).round(2)
    assert (aggregated["outstanding_amount"].round(2) == bucket_sum).all()

    # The grand total balance ties back to the canonical open-item total.
    assert round(float(total_row["outstanding_amount"]), 2) == round(
        float(document.data["outstanding_amount"].sum()), 2
    )


def test_ar_summary_pivot_uses_customer_grain():
    company, period = company_and_period()
    document = ARAgingGenerator().generate(company=company, period=period, seed=11)

    frame = build_report_frame(
        document=document,
        company=company,
        period=period,
        sheet_name="AR Aging",
        seed=2,
        severity=5,
    )

    assert frame.report_form == "summary_pivot"
    assert frame.grouping_key == "customer_id"
    assert "customer_name" in frame.aggregated_canonical_view.columns
    assert len(frame.aggregated_canonical_view) == document.data["customer_id"].nunique()


def test_low_severity_ap_stays_flat_listing():
    company, period = company_and_period()
    document = _ap_document()

    frame = build_report_frame(
        document=document,
        company=company,
        period=period,
        sheet_name="AP Aging",
        seed=1,
        severity=2,
    )

    assert frame.report_form == "listing"
    assert tuple(frame.table.columns) == tuple(document.data.columns)
    assert len(frame.table) == len(document.data)


def test_plain_listing_document_stays_listing():
    company, period = company_and_period()
    _, documents = _generate_documents(company, period, seed=42)
    # Payroll summary is not pivoted, grouped, or totaled, so it stays a flat list.
    payroll_summary = next(d for d in documents if d.document_type.value == "payroll_summary")

    frame = build_report_frame(
        document=payroll_summary,
        company=company,
        period=period,
        sheet_name="Payroll Summary",
        seed=3,
        severity=5,
    )

    assert frame.report_form == "listing"
    assert tuple(frame.table.columns) == tuple(payroll_summary.data.columns)
    assert len(frame.table) == len(payroll_summary.data)


def test_report_title_reads_like_a_finance_report():
    company, period = company_and_period()
    document = _ap_document()

    frame = build_report_frame(
        document=document,
        company=company,
        period=period,
        sheet_name="AP Aging",
        seed=1,
        severity=5,
    )

    assert "as at" in frame.report_title
    assert "31/12/2025" in frame.report_title
    assert frame.report_title.split(" as at")[0] in {
        "Aged Creditors",
        "Creditor Listing",
        "AP Ageing Summary",
        "Supplier Aging",
    }
    # Header band carries context that backend fields used to occupy.
    canonical_fields = {line["canonical_field"] for line in (l.as_dict() for l in frame.header_band)}
    assert {"client_id", "period_end", "currency"}.issubset(canonical_fields)


def test_generated_workbook_records_report_frame_and_scores_at_report_grain():
    company, period = company_and_period()
    config = load_config("config/nightmare.yaml")

    generated = generate_pbc_workbook_with_ground_truth(company, period, config=config, seed=42)
    metadata = generated.ground_truth.as_dict()

    ap_sheet = next(s for s in metadata["sheets"] if s["document_type"] == "ap_aging")
    assert ap_sheet["report_form"] == "summary_pivot"
    assert ap_sheet["report_archetype_id"] == "ap_aged_creditors_summary"
    assert ap_sheet["report_grain_schema"]
    assert ap_sheet["expected_report_output"]
    assert ap_sheet["report_header_band"]
    assert ap_sheet["report_grand_total"] is True
    assert "as at" in ap_sheet["report_title"]

    # The visible sheet no longer exposes the canonical backend spine.
    canonical_giveaways = {"client_id", "financial_year", "period_start", "period_end", "currency"}
    assert not (set(ap_sheet["visible_column_headers"]) & canonical_giveaways)

    # The recorded bucket labels are the labels actually rendered (no false claims).
    canonical_to_visible_all = ap_sheet["canonical_to_visible_columns"]
    for field, label in ap_sheet["report_bucket_label_mapping"].items():
        assert canonical_to_visible_all[field] == label

    # Simulate an extractor reading the visible pivot, then score it.
    canonical_to_visible = ap_sheet["canonical_to_visible_columns"]
    extraction_rows = [
        {canonical_to_visible[field]: value for field, value in row.items() if field in canonical_to_visible}
        for row in ap_sheet["expected_report_output"]
    ]
    extraction = {
        "documents": [
            {
                "document_type": "ap_aging",
                "sheet_name": "AP Aging",
                "table_location": ap_sheet["table_location"],
                "headers": ap_sheet["visible_column_headers"],
                "rows": extraction_rows,
            }
        ]
    }

    # Isolate the AP sheet so the single extraction document is matched to it
    # (the matcher otherwise greedily assigns one document to the first sheet).
    ap_only_groundtruth = dict(metadata)
    ap_only_groundtruth["sheets"] = [ap_sheet]
    report = compare_extraction(ap_only_groundtruth, extraction)
    ap_score = next(d for d in report.documents if d.document_type == "ap_aging")
    assert ap_score.metrics["column_mapping_accuracy"].score == 1.0
    assert ap_score.metrics["row_extraction_accuracy"].score == 1.0
    assert ap_score.metrics["numeric_value_accuracy"].score == 1.0
    # Phase 4: the score report surfaces the report form and the grain it scored at.
    assert ap_score.report_form == "summary_pivot"
    assert ap_score.scored_grain == "report"
    assert "Report form" in report.to_markdown()
    assert report.as_dict()["documents"][0]["report_form"] == "summary_pivot"


def _document(document_type, seed=42):
    company, period = company_and_period()
    _, documents = _generate_documents(company, period, seed=seed)
    return next(d for d in documents if d.document_type.value == document_type)


def test_gl_detail_grouped_inserts_subtotals_and_grand_total():
    company, period = company_and_period()
    document = _document("general_ledger")

    frame = build_report_frame(
        document=document,
        company=company,
        period=period,
        sheet_name="General Ledger",
        seed=1,
        severity=5,
    )

    assert frame.report_form == "detail_grouped"
    assert frame.grouping_key == "account_code"
    # Data grain is unchanged: aggregated view equals the canonical detail rows.
    assert len(frame.aggregated_canonical_view) == len(document.data)
    # The rendered table adds one subtotal per group plus a grand total.
    n_groups = document.data["account_code"].nunique()
    assert len(frame.table) == len(document.data) + n_groups + 1
    assert frame.subtotal_labels[-1] == "Grand Total"
    assert len(frame.subtotal_labels) == n_groups + 1

    # Grand total foots the amount columns and ties to the canonical detail.
    total_row = frame.table.iloc[-1]
    assert total_row["account_name"] == "Grand Total"
    for field in ("debit", "credit", "amount_signed"):
        assert round(float(total_row[field]), 2) == round(float(document.data[field].sum()), 2)


def test_grouped_detail_aggregated_view_excludes_total_rows():
    company, period = company_and_period()
    document = _document("general_ledger")

    frame = build_report_frame(
        document=document,
        company=company,
        period=period,
        sheet_name="General Ledger",
        seed=1,
        severity=5,
    )

    names = frame.aggregated_canonical_view["account_name"].astype(str)
    assert not names.str.startswith("Total").any()
    assert not (names == "Grand Total").any()


def test_payroll_detail_subtotals_labelled_by_department():
    company, period = company_and_period()
    document = _document("payroll_detail")

    frame = build_report_frame(
        document=document,
        company=company,
        period=period,
        sheet_name="Payroll Detail",
        seed=1,
        severity=5,
    )

    assert frame.report_form == "detail_grouped"
    assert frame.grouping_key == "department"
    displays = {label.removeprefix("Total - ") for label in frame.subtotal_labels if label != "Grand Total"}
    departments = {str(value) for value in document.data["department"].unique()}
    assert displays <= departments


def test_je_subtotals_labelled_by_journal_id():
    company, period = company_and_period()
    document = _document("journal_entry_listing")

    frame = build_report_frame(
        document=document,
        company=company,
        period=period,
        sheet_name="Journal Entries",
        seed=1,
        severity=5,
    )

    displays = {label.removeprefix("Total - ") for label in frame.subtotal_labels if label != "Grand Total"}
    journal_ids = {str(value) for value in document.data["journal_id"].unique()}
    assert displays <= journal_ids


def test_low_severity_gl_stays_flat_listing():
    company, period = company_and_period()
    document = _document("general_ledger")

    frame = build_report_frame(
        document=document,
        company=company,
        period=period,
        sheet_name="General Ledger",
        seed=1,
        severity=2,
    )

    assert frame.report_form == "listing"
    assert tuple(frame.table.columns) == tuple(document.data.columns)
    assert len(frame.table) == len(document.data)


def test_generated_grouped_workbook_scores_at_report_grain():
    company, period = company_and_period()
    config = load_config("config/nightmare.yaml")

    generated = generate_pbc_workbook_with_ground_truth(company, period, config=config, seed=42)
    metadata = generated.ground_truth.as_dict()
    gl_sheet = next(s for s in metadata["sheets"] if s["document_type"] == "general_ledger")

    assert gl_sheet["report_form"] == "detail_grouped"
    assert gl_sheet["report_subtotal_labels"]
    assert gl_sheet["report_subtotal_labels"][-1] == "Grand Total"

    canonical_to_visible = gl_sheet["canonical_to_visible_columns"]
    extraction_rows = [
        {canonical_to_visible[field]: value for field, value in row.items() if field in canonical_to_visible}
        for row in gl_sheet["expected_report_output"]
    ]
    extraction = {
        "documents": [
            {
                "document_type": "general_ledger",
                "sheet_name": "General Ledger",
                "table_location": gl_sheet["table_location"],
                "headers": gl_sheet["visible_column_headers"],
                "rows": extraction_rows,
            }
        ]
    }
    gl_only = dict(metadata)
    gl_only["sheets"] = [gl_sheet]
    report = compare_extraction(gl_only, extraction)
    gl_score = next(d for d in report.documents if d.document_type == "general_ledger")
    assert gl_score.metrics["column_mapping_accuracy"].score == 1.0
    assert gl_score.metrics["row_extraction_accuracy"].score == 1.0
    assert gl_score.metrics["numeric_value_accuracy"].score == 1.0


def test_generated_grouped_workbook_renders_grand_total():
    company, period = company_and_period()
    config = load_config("config/nightmare.yaml")
    generated = generate_pbc_workbook_with_ground_truth(company, period, config=config, seed=42)
    metadata = generated.ground_truth.as_dict()
    gl_sheet = next(s for s in metadata["sheets"] if s["document_type"] == "general_ledger")

    worksheet = generated.workbook[gl_sheet["sheet_name"]]
    cell_values = {
        str(cell.value)
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None
    }
    assert "Grand Total" in cell_values


def test_tb_totaled_listing_appends_grand_total():
    company, period = company_and_period()
    document = _document("trial_balance")

    frame = build_report_frame(
        document=document,
        company=company,
        period=period,
        sheet_name="Trial Balance",
        seed=1,
        severity=5,
    )

    assert frame.report_form == "totaled_listing"
    assert frame.grouping_key is None
    assert len(frame.aggregated_canonical_view) == len(document.data)
    assert len(frame.table) == len(document.data) + 1
    total_row = frame.table.iloc[-1]
    assert total_row["account_name"] == "Grand Total"
    for field in ("closing_debit", "closing_credit"):
        assert round(float(total_row[field]), 2) == round(float(document.data[field].sum()), 2)


def test_fixed_assets_grouped_by_class():
    company, period = company_and_period()
    document = _document("fixed_asset_register")

    frame = build_report_frame(
        document=document,
        company=company,
        period=period,
        sheet_name="Fixed Assets",
        seed=1,
        severity=5,
    )

    assert frame.report_form == "detail_grouped"
    assert frame.grouping_key == "asset_class"
    displays = {label.removeprefix("Total - ") for label in frame.subtotal_labels if label != "Grand Total"}
    assert displays <= {str(v) for v in document.data["asset_class"].unique()}


def test_inventory_grouped_by_category():
    company, period = company_and_period()
    document = _document("inventory_listing")

    frame = build_report_frame(
        document=document,
        company=company,
        period=period,
        sheet_name="Inventory",
        seed=1,
        severity=5,
    )

    assert frame.report_form == "detail_grouped"
    assert frame.grouping_key == "category"
    displays = {label.removeprefix("Total - ") for label in frame.subtotal_labels if label != "Grand Total"}
    assert displays <= {str(v) for v in document.data["category"].unique()}


def test_bank_recon_grouped_by_item_type():
    company, period = company_and_period()
    document = _document("bank_reconciliation")

    frame = build_report_frame(
        document=document,
        company=company,
        period=period,
        sheet_name="Bank Recon",
        seed=1,
        severity=5,
    )

    assert frame.report_form == "detail_grouped"
    assert frame.grouping_key == "recon_item_type"


def test_software_signature_is_deterministic_and_varied():
    signatures = {choose_software_signature(seed) for seed in range(20)}
    assert signatures <= set(SOFTWARE_SIGNATURES)
    assert len(signatures) >= 3
    assert choose_software_signature(5) == choose_software_signature(5)


def test_generated_workbook_records_and_renders_software_signature():
    company, period = company_and_period()
    generated = generate_pbc_workbook_with_ground_truth(
        company, period, config=load_config("config/nightmare.yaml"), seed=42
    )
    metadata = generated.ground_truth.as_dict()
    ap_sheet = next(s for s in metadata["sheets"] if s["document_type"] == "ap_aging")

    assert ap_sheet["report_software_signature"] in SOFTWARE_SIGNATURES

    worksheet = generated.workbook[ap_sheet["sheet_name"]]
    text_cells = [str(c.value) for row in worksheet.iter_rows() for c in row if isinstance(c.value, str)]
    assert any(ap_sheet["report_software_signature"] in value for value in text_cells)
    assert any(value == "Page 1 of 1" for value in text_cells)


def test_low_severity_workbook_has_no_software_signature():
    company, period = company_and_period()
    generated = generate_pbc_workbook_with_ground_truth(
        company, period, config=config_from_mapping({"severity": 2}), seed=42
    )
    metadata = generated.ground_truth.as_dict()
    for sheet in metadata["sheets"]:
        assert sheet["report_software_signature"] is None


def test_finance_number_formats_preserve_values():
    company, period = company_and_period()
    generated = generate_pbc_workbook_with_ground_truth(
        company, period, config=load_config("config/nightmare.yaml"), seed=42
    )
    metadata = generated.ground_truth.as_dict()
    ap_sheet = next(s for s in metadata["sheets"] if s["document_type"] == "ap_aging")
    worksheet = generated.workbook[ap_sheet["sheet_name"]]

    finance_formats = set(formatting._AMOUNT_FORMATS)
    numeric_with_finance_format = [
        cell
        for row in worksheet.iter_rows()
        for cell in row
        if isinstance(cell.value, (int, float))
        and not isinstance(cell.value, bool)
        and cell.number_format in finance_formats
    ]
    # Amount cells are formatted finance-style but remain real numbers (so scoring
    # still reads the underlying value, not formatted text).
    assert numeric_with_finance_format
    assert all(isinstance(cell.value, (int, float)) for cell in numeric_with_finance_format)


def test_generated_pivot_workbook_opens_and_has_grand_total():
    company, period = company_and_period()
    config = load_config("config/nightmare.yaml")

    generated = generate_pbc_workbook_with_ground_truth(company, period, config=config, seed=42)
    metadata = generated.ground_truth.as_dict()
    ap_sheet = next(s for s in metadata["sheets"] if s["document_type"] == "ap_aging")

    worksheet = generated.workbook[ap_sheet["sheet_name"]]
    cell_values = {
        str(cell.value)
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None
    }
    assert "Grand Total" in cell_values
