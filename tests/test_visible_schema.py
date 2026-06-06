from openpyxl import load_workbook

from pbc_chaos.config_loader import load_config
from pbc_chaos.generators.base import CompanyProfile, FinancialPeriod
from pbc_chaos.pbc_workbook import _generate_documents, generate_pbc_workbook_with_ground_truth
from pbc_chaos.scoring import compare_extraction
from pbc_chaos.workbook.visible_schema import COMMON_CONTEXT_FIELDS, build_visible_export


def company_and_period():
    return (
        CompanyProfile(company_id="client_001", company_name="ABC Sdn Bhd"),
        FinancialPeriod.calendar_year(2025),
    )


def documents_by_type(seed=42):
    company, period = company_and_period()
    _, documents = _generate_documents(company, period, seed=seed)
    return {document.document_type.value: document for document in documents}


def test_visible_export_profiles_drop_common_context_at_high_severity():
    company, period = company_and_period()
    document = documents_by_type()["trial_balance"]

    result = build_visible_export(
        document=document,
        company=company,
        period=period,
        sheet_name="Trial Balance",
        seed=42,
        severity=5,
    )

    assert "A/C" in result.data.columns
    assert "Description" in result.data.columns
    assert "client_id" not in result.data.columns
    assert set(COMMON_CONTEXT_FIELDS).issuperset(result.omitted_fields)
    assert result.context_field_values["financial_year"] == 2025
    assert result.visible_to_canonical["A/C"] == "account_code"
    assert result.visible_to_canonical["Description"] == "account_name"


def test_visible_export_keeps_canonical_baseline_for_severity_zero():
    company, period = company_and_period()
    document = documents_by_type()["trial_balance"]

    result = build_visible_export(
        document=document,
        company=company,
        period=period,
        sheet_name="Trial Balance",
        seed=42,
        severity=0,
    )

    assert tuple(result.data.columns) == tuple(document.data.columns)
    assert result.omitted_fields == ()
    assert result.visible_to_canonical["account_code"] == "account_code"


def test_visible_export_allows_same_header_to_mean_different_fields():
    company, period = company_and_period()
    documents = documents_by_type()

    gl = build_visible_export(
        document=documents["general_ledger"],
        company=company,
        period=period,
        sheet_name="General Ledger",
        seed=42,
        severity=5,
    )
    fixed_assets = build_visible_export(
        document=documents["fixed_asset_register"],
        company=company,
        period=period,
        sheet_name="Fixed Assets",
        seed=42,
        severity=5,
    )
    expense_claims = build_visible_export(
        document=documents["expense_claim_listing"],
        company=company,
        period=period,
        sheet_name="Expense Claims",
        seed=42,
        severity=5,
    )

    assert gl.visible_to_canonical["AC"] == "account_code"
    assert fixed_assets.visible_to_canonical["AC"] == "asset_class"
    assert expense_claims.visible_to_canonical["AC"] == "approval_status"


def test_generated_nightmare_workbook_hides_canonical_schema_spine(tmp_path):
    company, period = company_and_period()
    config = load_config("config/nightmare.yaml")
    generated = generate_pbc_workbook_with_ground_truth(company, period, config=config, seed=42)
    workbook_path = tmp_path / "visible.xlsx"
    generated.workbook.save(workbook_path)
    workbook = load_workbook(workbook_path)
    metadata = generated.ground_truth.as_dict()
    canonical_giveaways = {
        "client_id",
        "financial_year",
        "period_start",
        "period_end",
        "currency",
        "account_code",
        "account_name",
    }

    checked = 0
    for sheet in metadata["sheets"]:
        if sheet["document_type"] == "pbc_request_list":
            continue
        worksheet = workbook[sheet["sheet_name"]]
        header_row = sheet["table_location"]["header_row"]
        start_col = sheet["table_location"]["start_column"]
        end_col = sheet["table_location"]["end_column"]
        headers = {
            str(worksheet.cell(header_row, col).value)
            for col in range(start_col, end_col + 1)
            if worksheet.cell(header_row, col).value is not None
        }
        assert not (headers & canonical_giveaways)
        assert sheet["visible_export_profile"]
        assert sheet["visible_columns_mapping"]
        assert sheet["omitted_canonical_fields"]
        checked += 1

    assert checked >= 10


def test_scoring_accepts_visible_headers_and_context_fields():
    rows = [
        {
            "client_id": "client_001",
            "financial_year": 2025,
            "period_start": "2025-01-01",
            "period_end": "2025-12-31",
            "currency": "MYR",
            "account_code": "1000",
            "account_name": "Cash",
            "closing_balance": 123.45,
        },
        {
            "client_id": "client_001",
            "financial_year": 2025,
            "period_start": "2025-01-01",
            "period_end": "2025-12-31",
            "currency": "MYR",
            "account_code": "2000",
            "account_name": "AP",
            "closing_balance": -50.0,
        },
    ]
    table_location = {
        "start_row": 3,
        "start_column": 2,
        "end_row": 5,
        "end_column": 4,
        "header_row": 3,
    }
    groundtruth = {
        "workbook_id": "wb_test",
        "company_name": "ABC Sdn Bhd",
        "document_types_included": ["trial_balance"],
        "sheet_names": ["Trial Balance"],
        "injected_discrepancies": [],
        "sheets": [
            {
                "sheet_name": "Trial Balance",
                "document_type": "trial_balance",
                "clean_canonical_schema": [
                    "client_id",
                    "financial_year",
                    "period_start",
                    "period_end",
                    "currency",
                    "account_code",
                    "account_name",
                    "closing_balance",
                ],
                "visible_table_schema": [
                    "account_code",
                    "account_name",
                    "closing_balance",
                ],
                "visible_column_headers": ["A/C", "Description", "CY Bal"],
                "visible_columns_mapping": {
                    "A/C": "account_code",
                    "Description": "account_name",
                    "CY Bal": "closing_balance",
                },
                "canonical_to_visible_columns": {
                    "account_code": "A/C",
                    "account_name": "Description",
                    "closing_balance": "CY Bal",
                },
                "omitted_canonical_fields": [
                    "client_id",
                    "financial_year",
                    "period_start",
                    "period_end",
                    "currency",
                ],
                "context_field_values": {
                    "client_id": "client_001",
                    "financial_year": 2025,
                    "period_start": "2025-01-01",
                    "period_end": "2025-12-31",
                    "currency": "MYR",
                },
                "original_clean_row_count": 2,
                "final_messy_row_count": 2,
                "table_location": table_location,
                "expected_extraction_output": rows,
            }
        ],
    }
    extraction = {
        "documents": [
            {
                "document_type": "trial_balance",
                "sheet_name": "Trial Balance",
                "table_location": table_location,
                "headers": ["A/C", "Description", "CY Bal"],
                "rows": [
                    {"A/C": "1000", "Description": "Cash", "CY Bal": 123.45},
                    {"A/C": "2000", "Description": "AP", "CY Bal": -50.0},
                ],
            }
        ]
    }

    report = compare_extraction(groundtruth, extraction)

    assert report.overall_score == 1.0
    assert report.summary_metrics["column_mapping_accuracy"].score == 1.0
    assert report.summary_metrics["row_extraction_accuracy"].score == 1.0
