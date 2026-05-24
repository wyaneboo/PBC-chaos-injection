import json

from openpyxl import load_workbook

from pbc_chaos.config_loader import load_config
from pbc_chaos.generators.base import CompanyProfile, FinancialPeriod
from pbc_chaos.metadata.exporter import export_pbc_workbook
from pbc_chaos.pbc_workbook import generate_pbc_workbook_with_ground_truth


def company_and_period():
    return (
        CompanyProfile(company_id="client_001", company_name="ABC Sdn Bhd"),
        FinancialPeriod.calendar_year(2025),
    )


def test_workbook_ground_truth_contains_scoring_fields():
    company, period = company_and_period()
    config = load_config("config/nightmare.yaml")

    generated = generate_pbc_workbook_with_ground_truth(company, period, config=config, seed=42)
    metadata = generated.ground_truth.as_dict()

    assert metadata["workbook_id"]
    assert metadata["company_id"] == company.company_id
    assert metadata["company_name"] == company.company_name
    assert metadata["financial_period"]["financial_year"] == 2025
    assert metadata["seed"] == 42
    assert metadata["chaos_level"]["severity"] == 5
    assert "trial_balance" in metadata["document_types_included"]
    assert "Trial Balance" in metadata["sheet_names"]
    assert "trial_balance" in metadata["clean_canonical_schemas"]
    assert metadata["injected_discrepancies"]
    assert metadata["intentional_errors"]
    assert metadata["expected_extraction_output"]["trial_balance"]

    trial_balance = next(
        sheet for sheet in metadata["sheets"] if sheet["document_type"] == "trial_balance"
    )
    assert trial_balance["original_clean_row_count"] > 0
    assert trial_balance["final_messy_row_count"] >= trial_balance["original_clean_row_count"]
    assert trial_balance["table_location"]["start_cell"]
    assert trial_balance["table_location"]["end_cell"]
    assert trial_balance["table_location"]["header_row"] >= 1
    assert trial_balance["renamed_columns_mapping"]
    assert trial_balance["hidden_rows"]
    assert trial_balance["hidden_columns"]
    assert trial_balance["merged_cell_ranges"]
    assert trial_balance["inserted_notes"]
    assert trial_balance["intentional_errors"]


def test_exporter_writes_workbook_and_matching_json(tmp_path):
    company, period = company_and_period()

    exported = export_pbc_workbook(
        company,
        period,
        output_dir=tmp_path,
        config="config/nightmare.yaml",
        seed=42,
    )

    assert exported.workbook_path.name == "ABC_Sdn_Bhd_PBC_2025_nightmare.xlsx"
    assert exported.ground_truth_path.name == "ABC_Sdn_Bhd_PBC_2025_nightmare.groundtruth.json"
    assert exported.workbook_path.exists()
    assert exported.ground_truth_path.exists()

    reopened = load_workbook(exported.workbook_path)
    metadata = json.loads(exported.ground_truth_path.read_text(encoding="utf-8"))
    assert reopened.sheetnames == metadata["sheet_names"]
    assert metadata["workbook_id"] == exported.ground_truth.workbook_id

