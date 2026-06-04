from random import Random

from openpyxl import Workbook, load_workbook

from pbc_chaos.workbook import layout_engine, workbook_mutations


def make_workbook():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "PBC"
    worksheet.append(["Account", "Description", "Debit", "Credit"])
    worksheet.append(["1000", "Cash", 100.0, 0.0])
    worksheet.append(["2000", "Payables", 0.0, 75.0])
    worksheet.append(["4000", "Revenue", 0.0, 250.0])
    worksheet.append(["5000", "Payroll", 125.0, 0.0])
    return workbook, worksheet


def test_apply_layout_chaos_shifts_table_and_adds_workbook_artifacts(tmp_path):
    workbook, worksheet = make_workbook()

    layout_engine.apply_layout_chaos(
        workbook=workbook,
        worksheet=worksheet,
        config={"client_name": "Test Client", "old_version_tab_count": 1},
        seed=42,
    )

    assert worksheet["A1"].value is None
    assert worksheet.freeze_panes is not None
    assert any(str(value).startswith("Test Client") for row in worksheet.iter_rows(values_only=True) for value in row if value)
    assert any(name.startswith("PBC_old_v") for name in workbook.sheetnames)
    assert workbook["_recon_working"].sheet_state == "hidden"

    output_path = tmp_path / "messy.xlsx"
    workbook.save(output_path)
    reopened = load_workbook(output_path)
    assert "PBC" in reopened.sheetnames


def test_insert_blank_rows_and_columns_expands_bounds():
    workbook, worksheet = make_workbook()
    table = workbook_mutations.find_used_range(worksheet)

    updated = workbook_mutations.insert_blank_rows_and_columns(
        worksheet,
        table,
        row_count=1,
        column_count=1,
        rng=__import__("random").Random(7),
    )

    assert updated.max_row == table.max_row + 1
    assert updated.max_col == table.max_col + 1


def test_column_aliases_vary_for_same_canonical_header_across_sheets():
    aliases = set()

    for seed in range(30):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = f"Sheet {seed}"
        worksheet.append(["account_code"])
        worksheet.append(["1000"])
        table = workbook_mutations.find_used_range(worksheet)

        mapping = workbook_mutations.rename_random_columns(
            worksheet,
            table,
            count=1,
            rng=Random(seed),
            sheet_name=worksheet.title,
        )
        aliases.add(mapping["account_code"])

    assert len(aliases) >= 8
    assert "A/C" in aliases or "GL Code" in aliases


def test_reviewer_comments_and_status_cells_are_added():
    workbook, worksheet = make_workbook()

    layout_engine.apply_layout_chaos(
        workbook=workbook,
        worksheet=worksheet,
        config={
            "reviewer_comment_count": 2,
            "add_old_version_tabs": False,
            "add_hidden_reconciliation_tabs": False,
        },
        seed=5,
    )

    comments = [cell.comment for row in worksheet.iter_rows() for cell in row if cell.comment]
    statuses = [
        cell.value
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value in {"Reviewed", "Pending support", "Needs follow-up", "Client updated"}
    ]
    assert len(comments) == 2
    assert statuses


def test_config_rejects_unknown_keys():
    try:
        layout_engine.coerce_config({"not_a_setting": True})
    except ValueError as error:
        assert "not_a_setting" in str(error)
    else:
        raise AssertionError("Expected ValueError for unknown config key")
