"""Batch generation, validation, and manifest helpers."""

from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path
from random import Random
from typing import Any, Callable, Iterable

from openpyxl import load_workbook

from pbc_chaos.chaos.engine import ChaosEngine
from pbc_chaos.config.settings import SimulatorSettings
from pbc_chaos.config_loader import ChaosWorkbookConfig, config_from_mapping
from pbc_chaos.financial_model.builder import FinancialModelBuilder
from pbc_chaos.generators.base import CompanyProfile, FinancialPeriod
from pbc_chaos.generators.registry import DocumentGeneratorRegistry
from pbc_chaos.metadata.exporter import ExportedGroundTruthWorkbook, export_pbc_workbook
from pbc_chaos.metadata.records import RunManifest
from pbc_chaos.metadata.writer import MetadataWriter
from pbc_chaos.validation.checks import ValidationIssue, ValidationReport
from pbc_chaos.workbook.renderer import WorkbookRenderer


MANIFEST_COLUMNS = (
    "workbook_id",
    "company_name",
    "period",
    "chaos_level",
    "file_path",
    "groundtruth_path",
    "document_types",
    "row_count",
    "discrepancy_count",
    "generated_at",
)

_PERIOD_RE = re.compile(r"^(?:FY)?(?P<year>\d{4})$", re.IGNORECASE)
_COMPANY_ADJECTIVES = (
    "Apex",
    "Bintang",
    "Cahaya",
    "Delta",
    "Evergreen",
    "Fajar",
    "Global",
    "Harapan",
    "Ikhlas",
    "Jaya",
    "Kencana",
    "Lestari",
)
_COMPANY_NOUNS = (
    "Manufacturing",
    "Trading",
    "Logistics",
    "Holdings",
    "Foods",
    "Retail",
    "Engineering",
    "Resources",
    "Technology",
    "Services",
    "Supplies",
    "Industries",
)


class BatchPipelineError(ValueError):
    """Raised for user-correctable batch CLI errors."""


@dataclass(frozen=True)
class GeneratedWorkbookRecord:
    """One generated workbook and its exported artifacts."""

    company: CompanyProfile
    period: FinancialPeriod
    chaos_level: int
    seed: int | None
    exported: ExportedGroundTruthWorkbook


@dataclass(frozen=True)
class BatchGenerationResult:
    """Result returned by batch and dataset generation."""

    records: tuple[GeneratedWorkbookRecord, ...]
    manifest_path: Path


@dataclass(frozen=True)
class BatchSimulationPipeline:
    settings: SimulatorSettings
    financial_model_builder: FinancialModelBuilder
    generator_registry: DocumentGeneratorRegistry
    chaos_engine: ChaosEngine
    renderer: WorkbookRenderer
    metadata_writer: MetadataWriter

    def run(self) -> RunManifest:
        """Run the legacy settings-based batch pipeline.

        The Phase 9 CLI uses the helper functions in this module directly. This
        method is kept for the architecture contract and now produces a concrete
        run manifest from the first configured financial year.
        """

        severity_by_name = {"low": 1, "medium": 3, "high": 5}
        chaos_level = severity_by_name[self.settings.chaos.severity.value]
        year = self.settings.batch.financial_years[0]
        result = generate_batch_workbooks(
            company_count=self.settings.batch.client_count,
            period_label=f"FY{year}",
            chaos_level=chaos_level,
            output_dir=self.settings.run.output_dir,
            seed=self.settings.run.seed,
        )
        return RunManifest(
            run_id=self.settings.run.run_name or f"run_{self.settings.run.seed}",
            seed=self.settings.run.seed,
            output_dir=self.settings.run.output_dir,
            summary={
                "workbook_count": len(result.records),
                "manifest_path": str(result.manifest_path),
                "workbooks": [
                    {
                        "workbook_id": record.exported.ground_truth.workbook_id,
                        "path": str(record.exported.workbook_path),
                        "company_id": record.company.company_id,
                        "company_name": record.company.company_name,
                        "financial_year": record.period.financial_year,
                        "chaos_level": record.chaos_level,
                        "seed": record.seed,
                    }
                    for record in result.records
                ],
            },
        )


def generate_single_workbook(
    *,
    company_name: str,
    period_label: str,
    chaos_level: int,
    output_dir: str | Path,
    seed: int | None = None,
    company_id: str | None = None,
    filename_stem: str | None = None,
    unreproducible_nightmare_mode: bool = False,
    nightmare_progress_callback: Callable[[dict[str, Any]], None] | None = None,
) -> BatchGenerationResult:
    """Generate one workbook, sidecar JSON, and a one-row manifest."""

    period = parse_period(period_label)
    config = config_from_chaos_level(
        chaos_level,
        unreproducible_nightmare_mode=unreproducible_nightmare_mode,
    )
    company = CompanyProfile(
        company_id=company_id or company_id_from_name(company_name),
        company_name=company_name.strip(),
    )
    if not company.company_name:
        raise BatchPipelineError("Company name cannot be empty.")

    exported = _export_one(
        company=company,
        period=period,
        config=config,
        chaos_level=chaos_level,
        output_dir=output_dir,
        seed=seed,
        filename_stem=filename_stem,
        nightmare_progress_callback=nightmare_progress_callback,
    )
    record = GeneratedWorkbookRecord(
        company=company,
        period=period,
        chaos_level=chaos_level,
        seed=seed,
        exported=exported,
    )
    manifest_path = write_manifest_rows(
        [_manifest_row_from_record(record)],
        output_path=Path(output_dir) / "manifest.csv",
    )
    return BatchGenerationResult(records=(record,), manifest_path=manifest_path)


def generate_batch_workbooks(
    *,
    company_count: int,
    period_label: str,
    chaos_level: int,
    output_dir: str | Path,
    seed: int = 1,
    unreproducible_nightmare_mode: bool = False,
) -> BatchGenerationResult:
    """Generate many workbooks at one chaos level."""

    period = parse_period(period_label)
    config = config_from_chaos_level(
        chaos_level,
        unreproducible_nightmare_mode=unreproducible_nightmare_mode,
    )
    companies = simulated_companies(company_count, seed=seed)
    records = tuple(
        _generate_company_record(
            company=company,
            period=period,
            config=config,
            chaos_level=chaos_level,
            output_dir=output_dir,
            seed=seed + index,
            index=index,
        )
        for index, company in enumerate(companies, start=1)
    )
    manifest_path = write_manifest_rows(
        [_manifest_row_from_record(record) for record in records],
        output_path=Path(output_dir) / "manifest.csv",
    )
    return BatchGenerationResult(records=records, manifest_path=manifest_path)


def generate_mixed_chaos_dataset(
    *,
    company_count: int,
    period_label: str,
    min_chaos: int,
    max_chaos: int,
    output_dir: str | Path,
    seed: int = 1,
    unreproducible_nightmare_mode: bool = False,
) -> BatchGenerationResult:
    """Generate a dataset with chaos levels distributed across a range."""

    _validate_company_count(company_count)
    min_level = validate_chaos_level(min_chaos, label="min-chaos")
    max_level = validate_chaos_level(max_chaos, label="max-chaos")
    if min_level > max_level:
        raise BatchPipelineError("min-chaos must be less than or equal to max-chaos.")

    period = parse_period(period_label)
    levels = tuple(range(min_level, max_level + 1))
    companies = simulated_companies(company_count, seed=seed)
    rng = Random(seed)
    offset = rng.randrange(len(levels))
    records = []
    for index, company in enumerate(companies, start=1):
        chaos_level = levels[(index - 1 + offset) % len(levels)]
        records.append(
            _generate_company_record(
                company=company,
                period=period,
                config=config_from_chaos_level(
                    chaos_level,
                    unreproducible_nightmare_mode=unreproducible_nightmare_mode,
                ),
                chaos_level=chaos_level,
                output_dir=output_dir,
                seed=seed + index,
                index=index,
            )
        )
    manifest_path = write_manifest_rows(
        [_manifest_row_from_record(record) for record in records],
        output_path=Path(output_dir) / "manifest.csv",
    )
    return BatchGenerationResult(records=tuple(records), manifest_path=manifest_path)


def export_manifest(*, input_dir: str | Path, output_path: str | Path) -> Path:
    """Build and write a dataset manifest CSV from generated sidecar files."""

    input_path = _resolve_input_dir(input_dir)
    rows = build_manifest_rows(input_path)
    if not rows:
        raise BatchPipelineError(f"No generated workbooks found in {input_path}.")
    return write_manifest_rows(rows, output_path=output_path)


def write_manifest_rows(rows: Iterable[dict[str, str]], *, output_path: str | Path) -> Path:
    """Write manifest rows to CSV using the public Phase 9 schema."""

    row_list = list(rows)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=MANIFEST_COLUMNS)
        writer.writeheader()
        writer.writerows(row_list)
    return output


def build_manifest_rows(input_dir: str | Path) -> list[dict[str, str]]:
    """Return manifest rows derived from `.groundtruth.json` sidecars."""

    input_path = _resolve_input_dir(input_dir)
    rows = []
    for groundtruth_path in sorted(input_path.glob("*.groundtruth.json")):
        workbook_path = _workbook_path_for_groundtruth(groundtruth_path)
        if not workbook_path.exists():
            raise BatchPipelineError(
                f"Missing workbook for ground-truth sidecar: {groundtruth_path}"
            )
        ground_truth = _read_ground_truth(groundtruth_path)
        rows.append(_manifest_row(workbook_path, groundtruth_path, ground_truth))
    return rows


def validate_generated_directory(input_dir: str | Path) -> ValidationReport:
    """Validate generated workbook files and ground-truth sidecars."""

    input_path = Path(input_dir)
    issues: list[ValidationIssue] = []
    if not input_path.exists():
        issues.append(
            ValidationIssue(
                code="input_missing",
                message=f"Input directory does not exist: {input_path}",
                path=input_path,
            )
        )
        return ValidationReport(run_path=input_path, issues=tuple(issues))
    if not input_path.is_dir():
        issues.append(
            ValidationIssue(
                code="input_not_directory",
                message=f"Input path is not a directory: {input_path}",
                path=input_path,
            )
        )
        return ValidationReport(run_path=input_path, issues=tuple(issues))

    workbook_paths = sorted(path for path in input_path.glob("*.xlsx") if not path.name.startswith("~$"))
    groundtruth_paths = sorted(input_path.glob("*.groundtruth.json"))
    if not workbook_paths:
        issues.append(
            ValidationIssue(
                code="no_workbooks",
                message=f"No .xlsx workbooks found in {input_path}.",
                path=input_path,
            )
        )

    expected_sidecars = {_groundtruth_path_for_workbook(path) for path in workbook_paths}
    for workbook_path in workbook_paths:
        sidecar_path = _groundtruth_path_for_workbook(workbook_path)
        if not sidecar_path.exists():
            issues.append(
                ValidationIssue(
                    code="missing_groundtruth",
                    message=f"Missing ground-truth sidecar for {workbook_path.name}.",
                    path=workbook_path,
                )
            )
            continue
        _validate_workbook_pair(workbook_path, sidecar_path, issues)

    for sidecar_path in groundtruth_paths:
        if sidecar_path not in expected_sidecars:
            issues.append(
                ValidationIssue(
                    code="orphan_groundtruth",
                    message=f"Ground-truth sidecar has no matching workbook: {sidecar_path.name}.",
                    path=sidecar_path,
                )
            )

    return ValidationReport(run_path=input_path, issues=tuple(issues))


def parse_period(period_label: str) -> FinancialPeriod:
    """Parse `FY2025` or `2025` into a calendar-year financial period."""

    match = _PERIOD_RE.match(period_label.strip())
    if not match:
        raise BatchPipelineError("Period must use FY2025 or 2025 format.")
    return FinancialPeriod.calendar_year(int(match.group("year")))


def config_from_chaos_level(
    chaos_level: int,
    *,
    unreproducible_nightmare_mode: bool = False,
) -> ChaosWorkbookConfig:
    """Build a workbook config from an integer chaos level."""

    raw: dict[str, Any] = {"severity": validate_chaos_level(chaos_level)}
    if unreproducible_nightmare_mode:
        raw["unreproducible_nightmare_mode"] = {"enabled": True}
    return config_from_mapping(raw)


def validate_chaos_level(chaos_level: int, *, label: str = "chaos-level") -> int:
    """Validate the public 0-5 chaos level range."""

    if isinstance(chaos_level, bool) or not isinstance(chaos_level, int):
        raise BatchPipelineError(f"{label} must be an integer from 0 to 5.")
    if not 0 <= chaos_level <= 5:
        raise BatchPipelineError(f"{label} must be an integer from 0 to 5.")
    return chaos_level


def company_id_from_name(company_name: str) -> str:
    """Build a stable company identifier from a display name."""

    slug = _slug(company_name).lower()
    return f"client_{slug}" if slug else "client_company"


def simulated_companies(company_count: int, *, seed: int) -> tuple[CompanyProfile, ...]:
    """Create deterministic simulated company profiles for batch generation."""

    _validate_company_count(company_count)
    rng = Random(seed)
    companies = []
    for index in range(1, company_count + 1):
        adjective = rng.choice(_COMPANY_ADJECTIVES)
        noun = rng.choice(_COMPANY_NOUNS)
        company_name = f"{adjective} {noun} {index:03d} Sdn Bhd"
        companies.append(
            CompanyProfile(
                company_id=f"client_{index:04d}",
                company_name=company_name,
            )
        )
    return tuple(companies)


def _generate_company_record(
    *,
    company: CompanyProfile,
    period: FinancialPeriod,
    config: ChaosWorkbookConfig,
    chaos_level: int,
    output_dir: str | Path,
    seed: int,
    index: int,
) -> GeneratedWorkbookRecord:
    exported = _export_one(
        company=company,
        period=period,
        config=config,
        chaos_level=chaos_level,
        output_dir=output_dir,
        seed=seed,
        filename_stem=_filename_stem(company, period, chaos_level, index),
    )
    return GeneratedWorkbookRecord(
        company=company,
        period=period,
        chaos_level=chaos_level,
        seed=seed,
        exported=exported,
    )


def _export_one(
    *,
    company: CompanyProfile,
    period: FinancialPeriod,
    config: ChaosWorkbookConfig,
    chaos_level: int,
    output_dir: str | Path,
    seed: int | None,
    filename_stem: str | None,
    nightmare_progress_callback: Callable[[dict[str, Any]], None] | None = None,
) -> ExportedGroundTruthWorkbook:
    return export_pbc_workbook(
        company,
        period,
        output_dir=output_dir,
        config=config,
        seed=seed,
        filename_stem=filename_stem,
        nightmare_progress_callback=nightmare_progress_callback,
    )


def _manifest_row(
    workbook_path: Path,
    groundtruth_path: Path,
    ground_truth: dict[str, Any],
) -> dict[str, str]:
    period = ground_truth.get("financial_period", {})
    chaos_level = ground_truth.get("chaos_level", {})
    sheets = ground_truth.get("sheets", [])
    document_types = ground_truth.get("document_types_included", [])
    row_count = sum(_sheet_row_count(sheet) for sheet in sheets if isinstance(sheet, dict))
    return {
        "workbook_id": str(ground_truth.get("workbook_id", "")),
        "company_name": str(ground_truth.get("company_name", "")),
        "period": f"FY{period.get('financial_year', '')}",
        "chaos_level": str(chaos_level.get("severity", "")),
        "file_path": str(workbook_path),
        "groundtruth_path": str(groundtruth_path),
        "document_types": ";".join(str(document_type) for document_type in document_types),
        "row_count": str(row_count),
        "discrepancy_count": str(len(ground_truth.get("injected_discrepancies", []))),
        "generated_at": str(ground_truth.get("generated_at", "")),
    }


def _manifest_row_from_record(record: GeneratedWorkbookRecord) -> dict[str, str]:
    return _manifest_row(
        record.exported.workbook_path,
        record.exported.ground_truth_path,
        record.exported.ground_truth.as_dict(),
    )


def _validate_workbook_pair(
    workbook_path: Path,
    groundtruth_path: Path,
    issues: list[ValidationIssue],
) -> None:
    try:
        ground_truth = _read_ground_truth(groundtruth_path)
    except BatchPipelineError as exc:
        issues.append(
            ValidationIssue(
                code="invalid_groundtruth",
                message=str(exc),
                path=groundtruth_path,
            )
        )
        return

    for key in ("workbook_id", "company_name", "financial_period", "chaos_level", "sheets"):
        if key not in ground_truth:
            issues.append(
                ValidationIssue(
                    code="groundtruth_missing_key",
                    message=f"Ground-truth sidecar is missing required key: {key}.",
                    path=groundtruth_path,
                )
            )

    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    except Exception as exc:  # pragma: no cover - depends on openpyxl exception subclasses.
        issues.append(
            ValidationIssue(
                code="invalid_workbook",
                message=f"Could not open workbook {workbook_path.name}: {exc}",
                path=workbook_path,
            )
        )
        return

    expected_sheet_names = tuple(ground_truth.get("sheet_names", ()))
    if expected_sheet_names and tuple(workbook.sheetnames) != expected_sheet_names:
        issues.append(
            ValidationIssue(
                code="sheet_names_mismatch",
                message=f"Workbook sheets do not match sidecar for {workbook_path.name}.",
                path=workbook_path,
            )
        )

    for sheet in ground_truth.get("sheets", []):
        if not isinstance(sheet, dict):
            issues.append(
                ValidationIssue(
                    code="invalid_sheet_groundtruth",
                    message="A sheet ground-truth entry is not a JSON object.",
                    path=groundtruth_path,
                )
            )
            continue
        for key in ("sheet_name", "document_type", "table_location", "original_clean_row_count"):
            if key not in sheet:
                issues.append(
                    ValidationIssue(
                        code="sheet_groundtruth_missing_key",
                        message=f"Sheet ground truth is missing required key: {key}.",
                        path=groundtruth_path,
                    )
                )

    workbook.close()


def _read_ground_truth(path: Path) -> dict[str, Any]:
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise BatchPipelineError(f"Invalid JSON in {path}: {exc}") from exc
    if not isinstance(raw, dict):
        raise BatchPipelineError(f"Ground-truth sidecar must contain a JSON object: {path}")
    return raw


def _resolve_input_dir(input_dir: str | Path) -> Path:
    input_path = Path(input_dir)
    if not input_path.exists():
        raise BatchPipelineError(f"Input directory does not exist: {input_path}")
    if not input_path.is_dir():
        raise BatchPipelineError(f"Input path is not a directory: {input_path}")
    return input_path


def _validate_company_count(company_count: int) -> None:
    if isinstance(company_count, bool) or not isinstance(company_count, int) or company_count < 1:
        raise BatchPipelineError("companies must be a positive integer.")


def _sheet_row_count(sheet: dict[str, Any]) -> int:
    value = sheet.get("original_clean_row_count", 0)
    if isinstance(value, bool):
        return 0
    if isinstance(value, int | float):
        return int(value)
    return 0


def _workbook_path_for_groundtruth(groundtruth_path: Path) -> Path:
    return groundtruth_path.with_name(groundtruth_path.name.removesuffix(".groundtruth.json") + ".xlsx")


def _groundtruth_path_for_workbook(workbook_path: Path) -> Path:
    return workbook_path.with_suffix(".groundtruth.json")


def _filename_stem(
    company: CompanyProfile,
    period: FinancialPeriod,
    chaos_level: int,
    index: int,
) -> str:
    return (
        f"{index:04d}_{_slug(company.company_name)}_PBC_"
        f"{period.financial_year}_chaos_{chaos_level}"
    )


def _slug(value: object) -> str:
    text = re.sub(r"[^A-Za-z0-9]+", "_", str(value)).strip("_")
    return text or "workbook"
