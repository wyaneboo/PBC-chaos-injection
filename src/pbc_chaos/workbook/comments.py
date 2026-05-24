"""Reviewer comment mutations for PBC Excel workbooks."""

from __future__ import annotations

from random import Random

from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet

from pbc_chaos.workbook.workbook_mutations import TableBounds


COMMENT_TEXTS: tuple[str, ...] = (
    "Please tie this amount to the latest GL export.",
    "Client to confirm whether this includes late manual journals.",
    "Support received, but naming does not match the request list.",
    "Need explanation for movement versus prior year.",
    "Marked for follow-up during fieldwork.",
)


def add_reviewer_comments(
    worksheet: Worksheet,
    table: TableBounds,
    *,
    reviewer_name: str,
    count: int,
    rng: Random,
) -> None:
    """Attach Excel comments to a few non-header cells."""

    if table.max_row <= table.min_row:
        return

    candidates = [
        (row, col)
        for row in range(table.min_row + 1, table.max_row + 1)
        for col in range(table.min_col, table.max_col + 1)
        if worksheet.cell(row, col).value is not None
    ]
    rng.shuffle(candidates)

    for row, col in candidates[: max(0, count)]:
        worksheet.cell(row, col).comment = Comment(rng.choice(COMMENT_TEXTS), reviewer_name)

