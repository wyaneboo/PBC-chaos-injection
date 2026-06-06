"""Formatting mutations for manually maintained audit PBC workbooks."""

from __future__ import annotations

from datetime import date, datetime
from random import Random

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from pbc_chaos.workbook.workbook_mutations import TableBounds


#: Finance amount number formats: thousands, parenthesised negatives, dash zero,
#: and occasionally an inline RM currency prefix.
_AMOUNT_FORMATS: tuple[str, ...] = (
    '#,##0.00;(#,##0.00);"-"',
    "#,##0.00;(#,##0.00)",
    '"RM" #,##0.00;("RM" #,##0.00);"-"',
    '#,##0.00;[Red](#,##0.00);"-"',
)

#: Finance date formats; sheets mix a primary and a secondary format.
_DATE_FORMATS: tuple[str, ...] = ("dd/mm/yyyy", "dd-mmm-yy", "dd.mm.yyyy", "dd/mm/yy")


STATUS_STYLES: dict[str, str] = {
    "Reviewed": "C6E0B4",
    "Pending support": "FFF2CC",
    "Needs follow-up": "F8CBAD",
    "Client updated": "BDD7EE",
}


def apply_inconsistent_formatting(worksheet: Worksheet, table: TableBounds, *, rng: Random) -> None:
    """Apply small, plausible formatting inconsistencies across the table."""

    header_fill = PatternFill("solid", fgColor=rng.choice(["1F4E78", "5B9BD5", "70AD47"]))
    for col in range(table.min_col, table.max_col + 1):
        cell = worksheet.cell(table.header_row, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal=rng.choice(["center", "left"]))

    thin_gray = Side(style="thin", color="BFBFBF")
    dotted_gray = Side(style="hair", color="D9D9D9")
    for row in range(table.min_row + 1, table.max_row + 1):
        if row % rng.choice([5, 7, 9]) == 0:
            worksheet.row_dimensions[row].height = rng.choice([18, 21, 24])
        for col in range(table.min_col, table.max_col + 1):
            cell = worksheet.cell(row, col)
            cell.border = Border(bottom=rng.choice([thin_gray, dotted_gray]))
            if isinstance(cell.value, int | float):
                cell.number_format = rng.choice(["#,##0.00", "#,##0", "$#,##0.00"])
            if rng.random() < 0.035:
                cell.fill = PatternFill("solid", fgColor=rng.choice(["FFF2CC", "E2F0D9", "DDEBF7"]))
            if rng.random() < 0.02:
                cell.font = Font(italic=True, color="666666")

    for col in range(table.min_col, table.max_col + 1):
        worksheet.column_dimensions[worksheet.cell(table.header_row, col).column_letter].width = rng.choice(
            [10, 12, 14, 18, 22]
        )


def apply_finance_value_formats(
    worksheet: Worksheet,
    table: TableBounds,
    *,
    rng: Random,
) -> tuple[str, str]:
    """Apply finance-native number and date formats to the table's data cells.

    Amount columns get thousands separators, parenthesised negatives and a dash
    for zero (sometimes with an inline ``RM`` prefix). Date columns get a finance
    date format, with a minority rendered in a second format for realistic
    inconsistency. Underlying cell values are left numeric/date so extraction and
    scoring are unaffected; only the display format changes.

    Returns the chosen ``(amount_format, date_format)`` for logging.
    """

    amount_format = rng.choice(_AMOUNT_FORMATS)
    primary_date = rng.choice(_DATE_FORMATS)
    secondary_date = rng.choice(_DATE_FORMATS)

    amount_columns = _amount_columns(worksheet, table)
    date_columns = _date_columns(worksheet, table)

    for col in amount_columns:
        for row in range(table.min_row + 1, table.max_row + 1):
            cell = worksheet.cell(row, col)
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                cell.number_format = amount_format
                cell.alignment = Alignment(horizontal="right")

    for col in date_columns:
        for row in range(table.min_row + 1, table.max_row + 1):
            cell = worksheet.cell(row, col)
            if isinstance(cell.value, (date, datetime)):
                cell.number_format = secondary_date if rng.random() < 0.15 else primary_date

    return amount_format, primary_date


def _amount_columns(worksheet: Worksheet, table: TableBounds) -> list[int]:
    """Columns that look like money: at least two numeric cells of real magnitude."""

    columns: list[int] = []
    for col in range(table.min_col, table.max_col + 1):
        magnitudes = [
            abs(float(worksheet.cell(row, col).value))
            for row in range(table.min_row + 1, table.max_row + 1)
            if isinstance(worksheet.cell(row, col).value, (int, float))
            and not isinstance(worksheet.cell(row, col).value, bool)
        ]
        if len(magnitudes) >= 2 and max(magnitudes) >= 50:
            columns.append(col)
    return columns


def _date_columns(worksheet: Worksheet, table: TableBounds) -> list[int]:
    columns: list[int] = []
    for col in range(table.min_col, table.max_col + 1):
        header = str(worksheet.cell(table.header_row, col).value or "").lower()
        if any(token in header for token in ("date", "dt", "as at", "as on")):
            columns.append(col)
            continue
        if any(
            isinstance(worksheet.cell(row, col).value, (date, datetime))
            for row in range(table.min_row + 1, min(table.max_row, table.min_row + 10) + 1)
        ):
            columns.append(col)
    return columns


def apply_status_cells(worksheet: Worksheet, table: TableBounds, *, rng: Random) -> TableBounds:
    """Add a color-coded audit status column adjacent to the table."""

    status_col = table.max_col + 1
    worksheet.cell(table.header_row, status_col, "Audit status")
    worksheet.cell(table.header_row, status_col).font = Font(bold=True, color="FFFFFF")
    worksheet.cell(table.header_row, status_col).fill = PatternFill("solid", fgColor="7F7F7F")

    statuses = tuple(STATUS_STYLES)
    for row in range(table.min_row + 1, table.max_row + 1):
        status = rng.choice(statuses)
        cell = worksheet.cell(row, status_col, status)
        cell.fill = PatternFill("solid", fgColor=STATUS_STYLES[status])
        cell.alignment = Alignment(horizontal="center")

    worksheet.column_dimensions[worksheet.cell(table.header_row, status_col).column_letter].width = 16
    return TableBounds(table.min_row, table.min_col, table.max_row, status_col)

