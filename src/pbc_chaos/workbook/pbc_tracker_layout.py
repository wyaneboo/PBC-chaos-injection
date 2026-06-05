"""PBC request tracker layout and workflow-specific chaos."""

from __future__ import annotations

from datetime import date, datetime
from random import Random, SystemRandom
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.worksheet import Worksheet

from pbc_chaos.workbook import workbook_mutations
from pbc_chaos.workbook.workbook_mutations import TableBounds


TRACKER_SHEET_NAME = "PBC Request List"

DISPLAY_HEADERS = {
    "request_number": "No.",
    "request_id": "Request\nID",
    "request_description": "Request /\nDescription",
    "detail_remark": "Details / Remark\n(if any)",
    "purpose": "Purpose",
    "period_label": "Period",
    "file_type_requested": "Format /\nFile Type",
    "owner_pic": "Owner\n(PIC)",
    "due_date": "Due Date",
    "status": "Status",
    "date_received": "Date\nReceived",
    "review_status": "Review\nStatus",
    "auditor_comment": "Auditor\nComment",
    "follow_up_required": "Follow Up??",
    "update_flag": "Update Flag",
    "remarks": "Remarks",
    "client_id": "Client ID",
    "financial_year": "FY",
}

COLUMN_WIDTHS = {
    "request_number": 6,
    "request_id": 10,
    "request_description": 20,
    "detail_remark": 18,
    "purpose": 16,
    "period_label": 14,
    "file_type_requested": 15,
    "owner_pic": 12,
    "due_date": 12,
    "status": 13,
    "date_received": 12,
    "review_status": 15,
    "auditor_comment": 17,
    "follow_up_required": 13,
    "update_flag": 10,
    "remarks": 14,
    "client_id": 12,
    "financial_year": 8,
}

HIDDEN_KEYS = {"update_flag", "remarks", "client_id", "financial_year"}
STATUS_VARIANTS = {
    "done": ("Done", "done", "received"),
    "received": ("received", "Received", "done"),
    "partial": ("partial", "Partial", "partial"),
    "in_progress": ("In progress", "in\nprogress", "In progress"),
    "not_started": ("not yet", "not started", "not yet"),
    "not_applicable": ("N/A", "-", "N/A"),
}
REVIEW_VARIANTS = ("ok", "OK", "pending", "Under review", "Rending", "-", "")
FOLLOW_UP_VARIANTS = {True: ("Y", "Y?", "Y (remind)", "Yes"), False: ("N", "-", "")}
FILE_TYPE_VARIANTS = (
    "Excel",
    "xlsx / csv\n(whatever u have)",
    "PDF",
    ".pdf",
    "Excel/PDF",
    "Word / Excel",
    "email copy",
    "Excel??",
    "xlsx",
)
COMMENT_POOL = (
    "Missing p.4-6",
    "Qty not tally",
    "Need aging analysis",
    "Hardcopy only so far",
    "Some balance diff, checking",
    "Blurry scan",
    "Y - See email 15/5",
    "Not provided",
)


def apply_tracker_layout_chaos(
    *,
    workbook: Workbook,
    worksheet: Worksheet,
    config: Any,
    rng: Random,
    metadata_logger: Any | None,
) -> Workbook:
    """Apply request-list-specific visual and workflow chaos."""

    table = workbook_mutations.find_used_range(worksheet)
    if table is None:
        raise ValueError("Cannot apply tracker chaos to an empty worksheet.")

    header_keys = _header_keys(worksheet, table)
    worksheet.insert_rows(1, 5)
    table = TableBounds(table.min_row + 5, table.min_col, table.max_row + 5, table.max_col)

    header_to_col = _header_to_col(worksheet, table, header_keys)
    if config.tracker_instruction_blocks:
        _add_instruction_blocks(worksheet, table, config)
        if metadata_logger is not None:
            metadata_logger.record_inserted_note(
                worksheet.title,
                kind="tracker_instruction_blocks",
                cell="A1",
                text="Client submission, PIC, format, and late-submission instruction blocks.",
            )

    renamed = _apply_display_headers(worksheet, table, header_keys)
    if metadata_logger is not None:
        metadata_logger.record_renamed_columns(worksheet.title, renamed)

    _format_tracker_grid(worksheet, table, header_keys)
    _normalize_tracker_values(worksheet, table, header_to_col)
    _apply_tracker_noise(worksheet, table, header_to_col, config, rng, metadata_logger)

    worksheet.freeze_panes = worksheet.cell(table.header_row + 1, table.min_col + 2).coordinate
    worksheet.auto_filter.ref = (
        f"{worksheet.cell(table.header_row, table.min_col).coordinate}:"
        f"{worksheet.cell(table.max_row, table.max_col).coordinate}"
    )

    if metadata_logger is not None:
        metadata_logger.set_table_location(worksheet.title, table, worksheet)

    if config.add_old_version_tabs:
        old_version_tabs = workbook_mutations.add_old_version_tabs(
            workbook,
            source=worksheet,
            count=config.old_version_tab_count,
            rng=rng,
        )
        if metadata_logger is not None and old_version_tabs:
            metadata_logger.record_workbook_artifact(
                artifact_type="old_version_tabs",
                details={"source_sheet": worksheet.title, "sheet_names": old_version_tabs},
            )

    return workbook


def is_tracker_sheet(worksheet: Worksheet) -> bool:
    return worksheet.title == TRACKER_SHEET_NAME


def apply_tracker_agent_action(
    *,
    worksheet: Worksheet,
    action: dict[str, Any],
    logger: Any,
    rng: SystemRandom,
) -> dict[str, Any] | None:
    """Apply one bounded nightmare-agent tracker action."""

    table = find_tracker_table(worksheet)
    if table is None:
        return None
    header_to_col = _display_header_to_col(worksheet, table)
    row = _select_tracker_row(worksheet, table, header_to_col, action, rng)
    if row is None:
        return None

    tool = str(action.get("tool", ""))
    row_key = worksheet.cell(row, header_to_col.get("request_id", table.min_col)).value

    if tool == "add_visible_tracker_comment":
        col = header_to_col.get("auditor_comment")
        if col is None:
            return None
        text = _bounded_text(action.get("text")) or rng.choice(COMMENT_POOL)
        cell = worksheet.cell(row, col, text)
        cell.font = Font(bold=True, color="C00000" if rng.random() < 0.5 else "7030A0")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        logger.record_inserted_note(
            worksheet.title,
            kind="ai_tracker_visible_comment",
            cell=cell.coordinate,
            text={"request_id": row_key, "text": text},
        )
        return {"tool": tool, "sheet": worksheet.title, "row": row, "request_id": row_key}

    if tool == "apply_tracker_status_variant":
        col = header_to_col.get("status")
        if col is None:
            return None
        status = _bounded_text(action.get("status")) or rng.choice(
            ("not yet", "in progress", "partial", "received", "Done")
        )
        cell = worksheet.cell(row, col, status)
        _style_status_cell(cell, status)
        logger.record_intentional_error(
            worksheet.title,
            error_type="ai_tracker_status_variant",
            details={"cell": cell.coordinate, "request_id": row_key, "new_value": status},
        )
        return {"tool": tool, "sheet": worksheet.title, "row": row, "request_id": row_key}

    if tool == "apply_tracker_deadline_noise":
        col = header_to_col.get("due_date")
        if col is None:
            return None
        value = rng.choice(("??", "???", "22-May", "15/5"))
        cell = worksheet.cell(row, col, value)
        cell.font = Font(bold=True, color="C00000")
        logger.record_intentional_error(
            worksheet.title,
            error_type="ai_tracker_deadline_noise",
            details={"cell": cell.coordinate, "request_id": row_key, "new_value": value},
        )
        return {"tool": tool, "sheet": worksheet.title, "row": row, "request_id": row_key}

    if tool == "highlight_tracker_update_row":
        _highlight_row(worksheet, table, row)
        logger.record_inserted_note(
            worksheet.title,
            kind="ai_tracker_update_highlight",
            cell=worksheet.cell(row, table.min_col).coordinate,
            text={"request_id": row_key},
        )
        return {"tool": tool, "sheet": worksheet.title, "row": row, "request_id": row_key}

    if tool == "apply_tracker_follow_up_noise":
        col = header_to_col.get("follow_up_required")
        if col is None:
            return None
        value = _bounded_text(action.get("style")) or rng.choice(("Y?", "Y (remind)", "-", "N"))
        cell = worksheet.cell(row, col, value)
        cell.font = Font(bold=True, color="7030A0" if "Y" in value else "000000")
        logger.record_intentional_error(
            worksheet.title,
            error_type="ai_tracker_follow_up_noise",
            details={"cell": cell.coordinate, "request_id": row_key, "new_value": value},
        )
        return {"tool": tool, "sheet": worksheet.title, "row": row, "request_id": row_key}

    return None


def find_tracker_table(worksheet: Worksheet) -> TableBounds | None:
    for row in range(1, min(20, worksheet.max_row) + 1):
        values = [worksheet.cell(row, col).value for col in range(1, worksheet.max_column + 1)]
        normalized = {_normalize_header(value) for value in values if value is not None}
        if {"request_id", "request_description"} & normalized and "status" in normalized:
            max_col = max(
                col
                for col in range(1, worksheet.max_column + 1)
                if worksheet.cell(row, col).value is not None
            )
            return TableBounds(row, 1, worksheet.max_row, max_col)
    return None


def _header_keys(worksheet: Worksheet, table: TableBounds) -> tuple[str, ...]:
    return tuple(
        str(worksheet.cell(table.header_row, col).value or "").strip()
        for col in range(table.min_col, table.max_col + 1)
    )


def _header_to_col(
    worksheet: Worksheet,
    table: TableBounds,
    header_keys: tuple[str, ...],
) -> dict[str, int]:
    return {
        key: col
        for key, col in zip(header_keys, range(table.min_col, table.max_col + 1), strict=True)
        if key
    }


def _display_header_to_col(worksheet: Worksheet, table: TableBounds) -> dict[str, int]:
    found: dict[str, int] = {}
    for col in range(table.min_col, table.max_col + 1):
        key = _normalize_header(worksheet.cell(table.header_row, col).value)
        if key:
            found[key] = col
    return found


def _apply_display_headers(
    worksheet: Worksheet,
    table: TableBounds,
    header_keys: tuple[str, ...],
) -> dict[str, str]:
    renamed: dict[str, str] = {}
    for key, col in zip(header_keys, range(table.min_col, table.max_col + 1), strict=True):
        display = DISPLAY_HEADERS.get(key, key.replace("_", " ").title())
        worksheet.cell(table.header_row, col, display)
        if display != key:
            renamed[key] = display.replace("\n", " ")
    return renamed


def _add_instruction_blocks(worksheet: Worksheet, table: TableBounds, config: Any) -> None:
    max_col = max(table.max_col, table.min_col + 13)
    for row in range(1, table.header_row):
        worksheet.row_dimensions[row].height = 22

    def merge(start_col: int, end_col: int, row: int, end_row: int | None = None) -> None:
        worksheet.merge_cells(
            start_row=row,
            start_column=start_col,
            end_row=end_row or row,
            end_column=min(end_col, max_col),
        )

    merge(1, 3, 1, 2)
    worksheet["A1"] = f"{config.client_name}\n(123456-T)"
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet["A1"].alignment = Alignment(wrap_text=True, vertical="center")

    merge(4, 8, 1)
    worksheet["D1"] = "**PREPARED BY CLIENT (PBC) LIST**"
    worksheet["D1"].font = Font(bold=True, size=13)
    worksheet["D1"].alignment = Alignment(horizontal="center")
    merge(4, 8, 2)
    worksheet["D2"] = f"As at 15/05/{config.financial_year or 2025} (v3 - updated)"
    worksheet["D2"].font = Font(bold=True, color="FF0000", size=12)
    worksheet["D2"].alignment = Alignment(horizontal="center")

    merge(9, 10, 1, 2)
    worksheet["I1"] = CellRichText(
        "Pls prepare & submit by\n",
        TextBlock(InlineFont(b=True, color="C00000"), "latest 23 May (Fri)"),
    )
    worksheet["I1"].fill = PatternFill("solid", fgColor="FFF2CC")
    worksheet["I1"].alignment = Alignment(wrap_text=True)

    merge(13, 14, 1, 2)
    worksheet["M1"] = "PIC:\nFinance Team!!!"
    worksheet["M1"].fill = PatternFill("solid", fgColor="FFF2CC")
    worksheet["M1"].font = Font(bold=True, color="C00000")
    worksheet["M1"].alignment = Alignment(wrap_text=True)

    worksheet["A3"] = "Note:"
    worksheet["A4"] = "Highlighted = New / Updated"
    worksheet["A4"].fill = PatternFill("solid", fgColor="FFFF00")

    merge(5, 7, 3, 4)
    worksheet["E3"] = "Any Qs, call / msg\nJason (012-xxx 8899)"
    worksheet["E3"].font = Font(bold=True, color="FF5733")
    worksheet["E3"].alignment = Alignment(horizontal="center", wrap_text=True)

    merge(10, 11, 3, 4)
    worksheet["J3"] = "Format\nneed follow\nour template"
    worksheet["J3"].font = Font(color="7030A0")
    worksheet["J3"].alignment = Alignment(wrap_text=True)

    merge(12, 14, 3, 4)
    worksheet["L3"] = "Late submission may\nimpact audit timeline"
    worksheet["L3"].font = Font(bold=True, color="C00000")
    worksheet["L3"].alignment = Alignment(horizontal="center", wrap_text=True)


def _format_tracker_grid(worksheet: Worksheet, table: TableBounds, header_keys: tuple[str, ...]) -> None:
    header_fill = PatternFill("solid", fgColor="DDD9C4")
    thin = Side(style="thin", color="BFBFBF")
    medium = Side(style="medium", color="7F7F7F")
    for col, key in zip(range(table.min_col, table.max_col + 1), header_keys, strict=True):
        column_letter = get_column_letter(col)
        width = COLUMN_WIDTHS.get(key, 12)
        dimension: ColumnDimension = worksheet.column_dimensions[column_letter]
        dimension.width = width
        if key in HIDDEN_KEYS:
            dimension.hidden = True

        cell = worksheet.cell(table.header_row, col)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.border = Border(top=medium, bottom=medium, left=thin, right=thin)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    worksheet.row_dimensions[table.header_row].height = 34
    for row in range(table.header_row + 1, table.max_row + 1):
        worksheet.row_dimensions[row].height = 26 if row % 3 else 34
        for col in range(table.min_col, table.max_col + 1):
            cell = worksheet.cell(row, col)
            cell.border = Border(bottom=Side(style="hair", color="D9D9D9"))
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def _normalize_tracker_values(
    worksheet: Worksheet,
    table: TableBounds,
    header_to_col: dict[str, int],
) -> None:
    for row in range(table.header_row + 1, table.max_row + 1):
        follow_col = header_to_col.get("follow_up_required")
        if follow_col is not None:
            cell = worksheet.cell(row, follow_col)
            cell.value = "Y" if bool(cell.value) else "N"

        update_col = header_to_col.get("update_flag")
        if update_col is not None:
            cell = worksheet.cell(row, update_col)
            cell.value = "Updated" if bool(cell.value) else ""

        status_col = header_to_col.get("status")
        if status_col is not None:
            cell = worksheet.cell(row, status_col)
            cell.value = _status_display(str(cell.value or ""))
            _style_status_cell(cell, str(cell.value or ""))

        for key in ("due_date", "date_received"):
            col = header_to_col.get(key)
            if col is not None:
                cell = worksheet.cell(row, col)
                if isinstance(cell.value, datetime | date):
                    cell.number_format = "d/m"


def _apply_tracker_noise(
    worksheet: Worksheet,
    table: TableBounds,
    header_to_col: dict[str, int],
    config: Any,
    rng: Random,
    metadata_logger: Any | None,
) -> None:
    for row in range(table.header_row + 1, table.max_row + 1):
        request_id = worksheet.cell(row, header_to_col.get("request_id", table.min_col)).value

        if config.tracker_status_noise:
            _apply_request_id_noise(worksheet, row, header_to_col, rng, metadata_logger, request_id)
            _apply_status_noise(worksheet, row, header_to_col, rng, metadata_logger, request_id)
            _apply_review_status_noise(worksheet, row, header_to_col, rng, metadata_logger, request_id)
            _apply_file_type_noise(worksheet, row, header_to_col, rng, metadata_logger, request_id)

        if config.tracker_deadline_noise:
            _apply_deadline_noise(worksheet, row, header_to_col, rng, metadata_logger, request_id)
            _apply_received_date_noise(worksheet, row, header_to_col, rng, metadata_logger, request_id)
            _apply_follow_up_noise(worksheet, row, header_to_col, rng, metadata_logger, request_id)

        if config.tracker_visible_comments:
            _apply_visible_comment_noise(worksheet, row, header_to_col, rng, metadata_logger, request_id)

        if config.tracker_update_highlights:
            update_col = header_to_col.get("update_flag")
            updated = bool(update_col and worksheet.cell(row, update_col).value)
            if updated or rng.random() < 0.12:
                _highlight_row(worksheet, table, row)
                if metadata_logger is not None:
                    metadata_logger.record_inserted_note(
                        worksheet.title,
                        kind="tracker_update_highlight",
                        cell=worksheet.cell(row, table.min_col).coordinate,
                        text={"request_id": request_id},
                    )


def _apply_request_id_noise(
    worksheet: Worksheet,
    row: int,
    header_to_col: dict[str, int],
    rng: Random,
    metadata_logger: Any | None,
    request_id: Any,
) -> None:
    col = header_to_col.get("request_id")
    if col is None or rng.random() > 0.45:
        return
    cell = worksheet.cell(row, col)
    number = str(cell.value or "").split(".")[-1].replace("-", "")
    variants = (f"A.{number}", f"A-{number}", f"A{number}", f"A . {number}")
    new_value = rng.choice(variants)
    if new_value == cell.value:
        return
    original = cell.value
    cell.value = new_value
    _record_tracker_error(
        metadata_logger,
        worksheet,
        "tracker_request_id_format_noise",
        cell,
        request_id,
        original,
        new_value,
    )


def _apply_status_noise(
    worksheet: Worksheet,
    row: int,
    header_to_col: dict[str, int],
    rng: Random,
    metadata_logger: Any | None,
    request_id: Any,
) -> None:
    col = header_to_col.get("status")
    if col is None:
        return
    cell = worksheet.cell(row, col)
    key = _status_key(str(cell.value or ""))
    new_value = rng.choice(STATUS_VARIANTS.get(key, ("-", "In progress", "not yet")))
    original = cell.value
    cell.value = new_value
    _style_status_cell(cell, new_value)
    _record_tracker_error(
        metadata_logger,
        worksheet,
        "tracker_status_variant_noise",
        cell,
        request_id,
        original,
        new_value,
    )


def _apply_review_status_noise(
    worksheet: Worksheet,
    row: int,
    header_to_col: dict[str, int],
    rng: Random,
    metadata_logger: Any | None,
    request_id: Any,
) -> None:
    col = header_to_col.get("review_status")
    if col is None:
        return
    cell = worksheet.cell(row, col)
    original = cell.value
    new_value = rng.choice(REVIEW_VARIANTS)
    cell.value = new_value
    cell.font = Font(color="7030A0", bold=new_value in {"Under review", "pending"})
    _record_tracker_error(
        metadata_logger,
        worksheet,
        "tracker_review_status_noise",
        cell,
        request_id,
        original,
        new_value,
    )


def _apply_file_type_noise(
    worksheet: Worksheet,
    row: int,
    header_to_col: dict[str, int],
    rng: Random,
    metadata_logger: Any | None,
    request_id: Any,
) -> None:
    col = header_to_col.get("file_type_requested")
    if col is None or rng.random() > 0.60:
        return
    cell = worksheet.cell(row, col)
    original = cell.value
    new_value = rng.choice(FILE_TYPE_VARIANTS)
    cell.value = new_value
    if rng.random() < 0.35:
        cell.font = Font(underline="single")
    _record_tracker_error(
        metadata_logger,
        worksheet,
        "tracker_file_type_noise",
        cell,
        request_id,
        original,
        new_value,
    )


def _apply_deadline_noise(
    worksheet: Worksheet,
    row: int,
    header_to_col: dict[str, int],
    rng: Random,
    metadata_logger: Any | None,
    request_id: Any,
) -> None:
    col = header_to_col.get("due_date")
    if col is None:
        return
    cell = worksheet.cell(row, col)
    original = cell.value
    if rng.random() < 0.22:
        new_value: Any = rng.choice(("??", "???", ""))
    elif isinstance(original, datetime | date):
        new_value = rng.choice(
            (
                f"{original.day}/{original.month}",
                original.strftime("%d-%b").lstrip("0"),
                original,
            )
        )
    else:
        new_value = rng.choice(("15/5", "16-May", "??"))
    cell.value = new_value
    if isinstance(new_value, str) and "?" in new_value:
        cell.font = Font(bold=True, color="C00000")
    _record_tracker_error(
        metadata_logger,
        worksheet,
        "tracker_deadline_noise",
        cell,
        request_id,
        original,
        new_value,
    )


def _apply_received_date_noise(
    worksheet: Worksheet,
    row: int,
    header_to_col: dict[str, int],
    rng: Random,
    metadata_logger: Any | None,
    request_id: Any,
) -> None:
    col = header_to_col.get("date_received")
    if col is None:
        return
    cell = worksheet.cell(row, col)
    original = cell.value
    if original is None or rng.random() < 0.30:
        new_value: Any = rng.choice(("-", "", "14/5", "16-May"))
    elif isinstance(original, datetime | date):
        new_value = rng.choice((f"{original.day}/{original.month}", original.strftime("%d-%b").lstrip("0")))
    else:
        new_value = original
    cell.value = new_value
    _record_tracker_error(
        metadata_logger,
        worksheet,
        "tracker_received_date_noise",
        cell,
        request_id,
        original,
        new_value,
    )


def _apply_follow_up_noise(
    worksheet: Worksheet,
    row: int,
    header_to_col: dict[str, int],
    rng: Random,
    metadata_logger: Any | None,
    request_id: Any,
) -> None:
    col = header_to_col.get("follow_up_required")
    if col is None:
        return
    cell = worksheet.cell(row, col)
    original = cell.value
    key = str(original).strip().upper().startswith("Y")
    new_value = rng.choice(FOLLOW_UP_VARIANTS[key])
    cell.value = new_value
    cell.font = Font(color="7030A0" if "Y" in str(new_value) else "000000", bold="Y" in str(new_value))
    _record_tracker_error(
        metadata_logger,
        worksheet,
        "tracker_follow_up_noise",
        cell,
        request_id,
        original,
        new_value,
    )


def _apply_visible_comment_noise(
    worksheet: Worksheet,
    row: int,
    header_to_col: dict[str, int],
    rng: Random,
    metadata_logger: Any | None,
    request_id: Any,
) -> None:
    col = header_to_col.get("auditor_comment")
    if col is None or rng.random() > 0.55:
        return
    cell = worksheet.cell(row, col)
    if cell.value in (None, "") or rng.random() < 0.45:
        cell.value = rng.choice(COMMENT_POOL)
    cell.font = Font(bold=rng.random() < 0.55, color=rng.choice(("000000", "C00000", "7030A0")))
    if rng.random() < 0.35:
        cell.comment = Comment(str(cell.value), "Audit")
    if metadata_logger is not None:
        metadata_logger.record_inserted_note(
            worksheet.title,
            kind="tracker_visible_comment",
            cell=cell.coordinate,
            text={"request_id": request_id, "text": cell.value},
        )


def _highlight_row(worksheet: Worksheet, table: TableBounds, row: int) -> None:
    for col in range(table.min_col, min(table.max_col, table.min_col + 13) + 1):
        worksheet.cell(row, col).fill = PatternFill("solid", fgColor="FFF2CC")


def _style_status_cell(cell, value: str) -> None:
    text = str(value or "").lower()
    if "done" in text or "received" in text:
        color = "008000"
    elif "not" in text or "?" in text:
        color = "FF0000"
    elif "partial" in text:
        color = "008000"
        cell.font = Font(color=color, underline="single")
        return
    elif "progress" in text:
        color = "006100"
    else:
        color = "000000"
    cell.font = Font(bold=True, color=color)


def _status_display(value: str) -> str:
    return {
        "done": "Done",
        "received": "received",
        "partial": "partial",
        "in_progress": "In progress",
        "not_started": "not started",
        "not_applicable": "N/A",
    }.get(value, value)


def _status_key(value: str) -> str:
    text = value.lower().replace(" ", "_").replace("\n", "_")
    if "progress" in text:
        return "in_progress"
    if "received" in text:
        return "received"
    if "partial" in text:
        return "partial"
    if "done" in text:
        return "done"
    if "n/a" in text or text == "-":
        return "not_applicable"
    if "not" in text:
        return "not_started"
    return text


def _record_tracker_error(
    metadata_logger: Any | None,
    worksheet: Worksheet,
    error_type: str,
    cell: Any,
    request_id: Any,
    original: Any,
    new_value: Any,
) -> None:
    if metadata_logger is None or original == new_value:
        return
    metadata_logger.record_intentional_error(
        worksheet.title,
        error_type=error_type,
        details={
            "cell": cell.coordinate,
            "request_id": request_id,
            "original_value": original,
            "new_value": new_value,
        },
    )


def _select_tracker_row(
    worksheet: Worksheet,
    table: TableBounds,
    header_to_col: dict[str, int],
    action: dict[str, Any],
    rng: SystemRandom,
) -> int | None:
    request_col = header_to_col.get("request_id")
    data_rows = list(range(table.header_row + 1, table.max_row + 1))
    if not data_rows:
        return None

    row_key = str(action.get("row_key", "")).strip()
    if row_key and request_col is not None:
        normalized_target = _normalize_row_key(row_key)
        for row in data_rows:
            value = worksheet.cell(row, request_col).value
            if _normalize_row_key(value) == normalized_target:
                return row
        return None

    return rng.choice(data_rows)


def _bounded_text(value: Any, *, max_length: int = 96) -> str:
    if not isinstance(value, str):
        return ""
    return value.strip()[:max_length]


def _normalize_row_key(value: Any) -> str:
    return "".join(ch for ch in str(value or "").upper() if ch.isalnum())


def _normalize_header(value: Any) -> str:
    text = str(value or "").lower().replace("\n", " ")
    replacements = {
        "no.": "request_number",
        "no": "request_number",
        "request id": "request_id",
        "request / description": "request_description",
        "request description": "request_description",
        "details / remark (if any)": "detail_remark",
        "details / remark": "detail_remark",
        "details remark": "detail_remark",
        "format / file type": "file_type_requested",
        "format file type": "file_type_requested",
        "owner (pic)": "owner_pic",
        "owner pic": "owner_pic",
        "due date": "due_date",
        "date received": "date_received",
        "review status": "review_status",
        "auditor comment": "auditor_comment",
        "follow up??": "follow_up_required",
        "follow up": "follow_up_required",
        "update flag": "update_flag",
        "client id": "client_id",
        "fy": "financial_year",
    }
    compact = " ".join(text.replace("?", "").split())
    return replacements.get(compact, compact.replace(" ", "_").replace("/", "_"))
