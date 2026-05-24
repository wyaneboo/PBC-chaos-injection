"""Structural workbook mutations for realistic PBC Excel layouts."""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from random import Random
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class TableBounds:
    """Bounds for the primary table after layout mutations."""

    min_row: int
    min_col: int
    max_row: int
    max_col: int

    @property
    def header_row(self) -> int:
        return self.min_row

    @property
    def width(self) -> int:
        return self.max_col - self.min_col + 1


def find_used_range(worksheet: Worksheet) -> TableBounds | None:
    """Return the rectangular range containing non-empty cells."""

    used_rows: list[int] = []
    used_cols: list[int] = []
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                used_rows.append(cell.row)
                used_cols.append(cell.column)

    if not used_rows or not used_cols:
        return None
    return TableBounds(min(used_rows), min(used_cols), max(used_rows), max(used_cols))


def shift_table_away_from_a1(
    worksheet: Worksheet,
    table: TableBounds,
    *,
    top_rows: int,
    left_columns: int,
) -> TableBounds:
    """Insert leading rows and columns so the table no longer begins at A1."""

    if top_rows > 0:
        worksheet.insert_rows(1, top_rows)
    if left_columns > 0:
        worksheet.insert_cols(1, left_columns)
    return TableBounds(
        table.min_row + top_rows,
        table.min_col + left_columns,
        table.max_row + top_rows,
        table.max_col + left_columns,
    )


def add_title_block(
    worksheet: Worksheet,
    table: TableBounds,
    *,
    title: str,
    client_name: str,
    prepared_by: str,
    financial_year: int | None,
    rng: Random,
    merge_cells: bool = True,
) -> TableBounds:
    """Add a finance-team style title and metadata block above the table."""

    insert_at = table.min_row
    worksheet.insert_rows(insert_at, 4)

    title_row = insert_at
    subtitle_row = insert_at + 1
    meta_row = insert_at + 2
    note_row = insert_at + 3
    min_col = table.min_col
    max_col = table.max_col

    if merge_cells:
        worksheet.merge_cells(
            start_row=title_row,
            start_column=min_col,
            end_row=title_row,
            end_column=max_col,
        )
    title_cell = worksheet.cell(title_row, min_col, title)
    title_cell.font = Font(bold=True, size=14, color="1F2937")
    title_cell.fill = PatternFill("solid", fgColor=rng.choice(["D9EAF7", "E2F0D9", "FFF2CC"]))
    title_cell.alignment = Alignment(horizontal="left")

    worksheet.cell(subtitle_row, min_col, f"Client: {client_name}")
    worksheet.cell(subtitle_row, min_col + 2, f"FY: {financial_year or 'per client file'}")
    worksheet.cell(meta_row, min_col, f"Prepared by: {prepared_by}")
    worksheet.cell(meta_row, min_col + 2, "Status: Draft for audit")
    worksheet.cell(note_row, min_col, "Please do not overwrite prior comments; audit team using this version.")

    return TableBounds(table.min_row + 4, table.min_col, table.max_row + 4, table.max_col)


def add_client_notes(worksheet: Worksheet, table: TableBounds, *, rng: Random) -> TableBounds:
    """Add short client/audit notes above the table."""

    notes = [
        "Note: balances per GL export before late adjustments.",
        "Client comment: pending final AP accrual entries.",
        "Audit request: add support references where available.",
    ]
    insert_at = table.min_row
    worksheet.insert_rows(insert_at, 2)
    worksheet.cell(insert_at, table.min_col, rng.choice(notes))
    worksheet.cell(insert_at + 1, table.min_col, "Updated manually by finance after close meeting.")
    return TableBounds(table.min_row + 2, table.min_col, table.max_row + 2, table.max_col)


def insert_blank_rows_and_columns(
    worksheet: Worksheet,
    table: TableBounds,
    *,
    row_count: int,
    column_count: int,
    rng: Random,
) -> TableBounds:
    """Insert blank separators in non-critical positions inside/near the table."""

    max_row = table.max_row
    max_col = table.max_col
    for _ in range(max(0, row_count)):
        if table.max_row - table.min_row < 3:
            break
        row_idx = rng.randint(table.min_row + 2, max_row)
        worksheet.insert_rows(row_idx, 1)
        max_row += 1

    for _ in range(max(0, column_count)):
        if table.max_col - table.min_col < 2:
            break
        col_idx = rng.randint(table.min_col + 1, max_col)
        worksheet.insert_cols(col_idx, 1)
        max_col += 1

    return TableBounds(table.min_row, table.min_col, max_row, max_col)


def duplicate_header_rows(
    worksheet: Worksheet,
    table: TableBounds,
    *,
    count: int,
    rng: Random,
) -> TableBounds:
    """Duplicate header rows where a user may have split a long table for viewing."""

    max_row = table.max_row
    for _ in range(max(0, count)):
        target = rng.randint(table.min_row + 3, max_row + 1)
        worksheet.insert_rows(target, 1)
        for col in range(table.min_col, table.max_col + 1):
            _copy_cell(worksheet.cell(table.header_row, col), worksheet.cell(target, col))
        max_row += 1
    return TableBounds(table.min_row, table.min_col, max_row, table.max_col)


def add_subtotal_rows(
    worksheet: Worksheet,
    table: TableBounds,
    *,
    count: int,
    rng: Random,
) -> TableBounds:
    """Insert subtotal rows with formulas under plausible table sections."""

    max_row = table.max_row
    numeric_cols = _numeric_columns(worksheet, table)
    for _ in range(max(0, count)):
        target = rng.randint(table.min_row + 4, max_row + 1)
        worksheet.insert_rows(target, 1)
        worksheet.cell(target, table.min_col, "Subtotal")
        worksheet.cell(target, table.min_col).font = Font(bold=True)
        for col in numeric_cols:
            letter = get_column_letter(col)
            start_row = max(table.min_row + 1, target - rng.randint(3, 8))
            worksheet.cell(target, col, f"=SUM({letter}{start_row}:{letter}{target - 1})")
            worksheet.cell(target, col).font = Font(bold=True)
        max_row += 1
    return TableBounds(table.min_row, table.min_col, max_row, table.max_col)


def add_footer_notes(
    worksheet: Worksheet,
    table: TableBounds,
    *,
    rng: Random,
    merge_cells: bool = True,
) -> TableBounds:
    """Add footer notes below the table, with occasional in-table reminder text."""

    note_row = table.max_row + 2
    worksheet.cell(note_row, table.min_col, rng.choice([
        "Footer note: excludes intercompany eliminations not posted at extract date.",
        "Prepared from ERP export; manual reclasses highlighted by finance.",
        "Older review notes intentionally retained for audit trail.",
    ]))
    if merge_cells:
        worksheet.merge_cells(
            start_row=note_row,
            start_column=table.min_col,
            end_row=note_row,
            end_column=table.max_col,
        )
    worksheet.cell(note_row, table.min_col).font = Font(italic=True, color="666666")

    if table.max_row - table.min_row > 5:
        reminder_row = rng.randint(table.min_row + 2, table.max_row - 1)
        worksheet.cell(reminder_row, table.max_col, worksheet.cell(reminder_row, table.max_col).value)
        worksheet.cell(reminder_row, table.max_col).fill = PatternFill("solid", fgColor="FFF2CC")

    return table


def freeze_panes_near_table(worksheet: Worksheet, table: TableBounds) -> None:
    """Freeze panes just below/right of the primary header."""

    worksheet.freeze_panes = worksheet.cell(table.header_row + 1, table.min_col + 1).coordinate


def hide_random_rows_and_columns(
    worksheet: Worksheet,
    table: TableBounds,
    *,
    row_count: int,
    column_count: int,
    rng: Random,
) -> None:
    """Hide non-header rows and non-key columns to mimic manual working files."""

    candidate_rows = list(range(table.min_row + 2, table.max_row + 1))
    rng.shuffle(candidate_rows)
    for row in candidate_rows[: max(0, row_count)]:
        worksheet.row_dimensions[row].hidden = True

    candidate_cols = list(range(table.min_col + 1, table.max_col + 1))
    rng.shuffle(candidate_cols)
    for col in candidate_cols[: max(0, column_count)]:
        worksheet.column_dimensions[get_column_letter(col)].hidden = True


def add_old_version_tabs(
    workbook: Workbook,
    *,
    source: Worksheet,
    count: int,
    rng: Random,
) -> None:
    """Copy the active worksheet into visible old-version tabs."""

    for idx in range(1, max(0, count) + 1):
        copied = workbook.copy_worksheet(source)
        copied.title = _safe_sheet_title(workbook, f"{source.title}_old_v{idx}")
        copied.sheet_properties.tabColor = rng.choice(["C9C9C9", "F4B183", "BDD7EE"])
        copied["A1"] = "OLD VERSION - retained for support trail"


def add_hidden_reconciliation_tab(workbook: Workbook, *, source: Worksheet) -> Worksheet:
    """Create a hidden working tab that looks like a client reconciliation scratchpad."""

    title = _safe_sheet_title(workbook, "_recon_working")
    sheet = workbook.create_sheet(title)
    sheet.sheet_state = "hidden"
    rows: tuple[tuple[Any, ...], ...] = (
        ("Source sheet", source.title),
        ("Check", "Amount per support"),
        ("GL total", None),
        ("PBC total", None),
        ("Variance", "=B3-B4"),
        ("Prepared by", "Client finance"),
    )
    for row in rows:
        sheet.append(row)
    sheet["A1"].font = Font(bold=True)
    return sheet


def rename_random_columns(
    worksheet: Worksheet,
    table: TableBounds,
    *,
    count: int,
    rng: Random,
) -> None:
    """Rename selected headers using common finance shorthand."""

    aliases = {
        "account_code": ("GL", "Acct No", "A/C"),
        "account_name": ("Account", "Description"),
        "closing_balance": ("Closing Bal", "Amount", "FY Bal"),
        "outstanding_amount": ("Outstanding", "Open Amt"),
        "posting_date": ("Post Dt", "GL Date"),
        "invoice_date": ("Inv Date", "Doc Date"),
        "vendor_name": ("Supplier", "Vendor"),
        "customer_name": ("Customer", "Debtor"),
        "debit": ("Dr", "Debit Amt"),
        "credit": ("Cr", "Credit Amt"),
        "amount_signed": ("Signed Amt", "Net Amt"),
        "gross_pay": ("Gross", "Gross Payroll"),
        "net_pay": ("Net", "Net Payroll"),
        "total_cost": ("Value", "Ext Cost"),
    }
    candidates = []
    for col in range(table.min_col, table.max_col + 1):
        value = worksheet.cell(table.header_row, col).value
        if value is not None:
            candidates.append((col, str(value)))
    rng.shuffle(candidates)
    for col, header in candidates[: max(0, count)]:
        key = header.strip().lower().replace(" ", "_")
        choices = aliases.get(key)
        worksheet.cell(table.header_row, col).value = rng.choice(choices) if choices else _humanize_header(header)


def stringify_numeric_cells(
    worksheet: Worksheet,
    table: TableBounds,
    *,
    count: int,
    rng: Random,
) -> None:
    """Convert selected numeric cells to text while preserving the visible value."""

    candidates = [
        cell
        for row in worksheet.iter_rows(
            min_row=table.min_row + 1,
            max_row=table.max_row,
            min_col=table.min_col,
            max_col=table.max_col,
        )
        for cell in row
        if isinstance(cell.value, int | float) and not isinstance(cell.value, bool)
    ]
    rng.shuffle(candidates)
    for cell in candidates[: max(0, count)]:
        cell.value = f"{float(cell.value):,.2f}"


def inject_formula_errors(
    worksheet: Worksheet,
    table: TableBounds,
    *,
    count: int,
    rng: Random,
) -> None:
    """Replace a few numeric cells with broken formulas."""

    candidates = [
        cell
        for row in worksheet.iter_rows(
            min_row=table.min_row + 1,
            max_row=table.max_row,
            min_col=table.min_col,
            max_col=table.max_col,
        )
        for cell in row
        if isinstance(cell.value, int | float) and not isinstance(cell.value, bool)
    ]
    rng.shuffle(candidates)
    errors = ("=#REF!", "=SUM(#REF!)", "=1/0")
    for cell in candidates[: max(0, count)]:
        cell.value = rng.choice(errors)


def insert_wrong_period_rows(
    worksheet: Worksheet,
    table: TableBounds,
    *,
    count: int,
    rng: Random,
) -> TableBounds:
    """Copy rows and move date/year values outside the reporting period."""

    date_columns = _date_like_columns(worksheet, table)
    if not date_columns or table.max_row <= table.min_row:
        return table

    max_row = table.max_row
    for _ in range(max(0, count)):
        source_row = rng.randint(table.min_row + 1, max_row)
        target_row = source_row + 1
        worksheet.insert_rows(target_row, 1)
        for col in range(table.min_col, table.max_col + 1):
            _copy_cell(worksheet.cell(source_row, col), worksheet.cell(target_row, col))
        for col in date_columns:
            cell = worksheet.cell(target_row, col)
            cell.value = _wrong_period_value(cell.value)
            cell.fill = PatternFill("solid", fgColor="F8CBAD")
        worksheet.cell(target_row, table.min_col).font = Font(italic=True, color="C00000")
        max_row += 1
    return TableBounds(table.min_row, table.min_col, max_row, table.max_col)


def add_secondary_tables(
    worksheet: Worksheet,
    table: TableBounds,
    *,
    count: int,
    rng: Random,
) -> None:
    """Add small side tables below the main table."""

    for index in range(max(0, count)):
        start_row = table.max_row + 4 + index * 6
        start_col = table.min_col + rng.randint(0, min(2, max(0, table.width - 3)))
        worksheet.cell(start_row, start_col, rng.choice(["Manual recon", "Finance notes", "Late adjustments"]))
        worksheet.cell(start_row, start_col).font = Font(bold=True)
        headers = ("Ref", "Description", "Amount")
        for offset, header in enumerate(headers):
            cell = worksheet.cell(start_row + 1, start_col + offset, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="7F7F7F")
        for row_offset in range(2, 5):
            worksheet.cell(start_row + row_offset, start_col, f"M-{index + 1}-{row_offset - 1}")
            worksheet.cell(
                start_row + row_offset,
                start_col + 1,
                rng.choice(["Accrual true-up", "Timing item", "Client estimate"]),
            )
            worksheet.cell(start_row + row_offset, start_col + 2, round(rng.uniform(-5000, 5000), 2))


def _numeric_columns(worksheet: Worksheet, table: TableBounds) -> list[int]:
    columns: list[int] = []
    for col in range(table.min_col, table.max_col + 1):
        numeric_cells = 0
        for row in range(table.min_row + 1, min(table.max_row, table.min_row + 10) + 1):
            if isinstance(worksheet.cell(row, col).value, int | float):
                numeric_cells += 1
        if numeric_cells >= 2:
            columns.append(col)
    return columns


def _date_like_columns(worksheet: Worksheet, table: TableBounds) -> list[int]:
    columns: list[int] = []
    for col in range(table.min_col, table.max_col + 1):
        header = str(worksheet.cell(table.header_row, col).value or "").lower()
        if any(token in header for token in ("date", "period", "year", "fy")):
            columns.append(col)
            continue
        for row in range(table.min_row + 1, min(table.max_row, table.min_row + 10) + 1):
            value = worksheet.cell(row, col).value
            if isinstance(value, date | datetime):
                columns.append(col)
                break
    return columns


def _wrong_period_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value + timedelta(days=370)
    if isinstance(value, date):
        return value + timedelta(days=370)
    if isinstance(value, int) and 1900 <= value <= 2200:
        return value + 1
    if isinstance(value, str):
        stripped = value.strip()
        if len(stripped) >= 4 and stripped[:4].isdigit():
            return f"{int(stripped[:4]) + 1}{stripped[4:]}"
    return value


def _humanize_header(header: str) -> str:
    words = header.replace("_", " ").split()
    if not words:
        return header
    return " ".join(word[:4].title() if len(word) > 4 else word.title() for word in words)


def _copy_cell(source: Cell, target: Cell) -> None:
    target.value = source.value
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def _safe_sheet_title(workbook: Workbook, title: str) -> str:
    base = title[:31]
    candidate = base
    suffix = 1
    while candidate in workbook.sheetnames:
        suffix_text = f"_{suffix}"
        candidate = f"{base[: 31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    return candidate
