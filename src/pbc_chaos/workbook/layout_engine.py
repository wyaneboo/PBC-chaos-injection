"""High-level Excel layout chaos engine.

This module mutates an already-populated openpyxl worksheet. It is intentionally
limited to workbook layout realism: no document generation and no reconciliation
logic are performed here.
"""

from __future__ import annotations

from dataclasses import dataclass, fields
from random import Random
from typing import Any, Mapping

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from pbc_chaos.workbook import pbc_tracker_layout
from pbc_chaos.workbook import comments, formatting, workbook_mutations


@dataclass(frozen=True)
class LayoutChaosConfig:
    """Controls the intensity and shape of Excel layout chaos.

    Defaults are calibrated for audit-season finance workbooks: tables are still
    usable, but the surrounding workbook has manual notes, old tabs, hidden
    review material, inconsistent styling, and small layout surprises.
    """

    client_name: str = "Example Client"
    prepared_by: str = "Finance Team"
    reviewer_name: str = "Audit reviewer"
    financial_year: int | None = None
    title: str | None = None
    enabled: bool = True
    max_top_shift: int = 5
    max_left_shift: int = 3
    blank_row_count: int = 2
    blank_column_count: int = 1
    merge_cells: bool = True
    duplicate_header_count: int = 1
    subtotal_row_count: int = 1
    hidden_row_count: int = 1
    hidden_column_count: int = 1
    reviewer_comment_count: int = 3
    add_title_block: bool = True
    add_client_notes_block: bool = True
    add_footer_notes_block: bool = True
    software_signature: str | None = None
    report_currency: str | None = None
    report_as_at: str | None = None
    rename_column_count: int = 0
    stringified_number_count: int = 0
    formula_error_count: int = 0
    wrong_period_row_count: int = 0
    secondary_table_count: int = 0
    add_old_version_tabs: bool = True
    old_version_tab_count: int = 1
    add_hidden_reconciliation_tabs: bool = True
    pbc_request_list_layout: bool = False
    tracker_status_noise: bool = False
    tracker_deadline_noise: bool = False
    tracker_visible_comments: bool = False
    tracker_update_highlights: bool = False
    tracker_instruction_blocks: bool = False


def coerce_config(config: LayoutChaosConfig | Mapping[str, Any] | None) -> LayoutChaosConfig:
    """Return a typed config from a dataclass, mapping, or ``None``."""

    if config is None:
        return LayoutChaosConfig()
    if isinstance(config, LayoutChaosConfig):
        return config

    allowed = {field.name for field in fields(LayoutChaosConfig)}
    unknown = set(config) - allowed
    if unknown:
        names = ", ".join(sorted(unknown))
        raise ValueError(f"Unknown layout chaos config keys: {names}")
    return LayoutChaosConfig(**dict(config))


def apply_layout_chaos(
    workbook: Workbook,
    worksheet: Worksheet,
    config: LayoutChaosConfig | Mapping[str, Any] | None = None,
    seed: int | None = None,
    metadata_logger: Any | None = None,
) -> Workbook:
    """Apply realistic Excel layout chaos to ``worksheet`` and return ``workbook``.

    The expected input is a clean worksheet containing one primary table with a
    header row at the top of the populated range. Mutations preserve workbook
    validity and aim to mimic common client-maintained PBC workbooks.
    """

    resolved = coerce_config(config)
    if not resolved.enabled:
        if metadata_logger is not None:
            table = workbook_mutations.find_used_range(worksheet)
            if table is not None:
                metadata_logger.set_table_location(worksheet.title, table, worksheet)
        return workbook

    rng = Random(seed)

    table = workbook_mutations.find_used_range(worksheet)
    if table is None:
        raise ValueError("Cannot apply layout chaos to an empty worksheet.")

    if resolved.pbc_request_list_layout and pbc_tracker_layout.is_tracker_sheet(worksheet):
        return pbc_tracker_layout.apply_tracker_layout_chaos(
            workbook=workbook,
            worksheet=worksheet,
            config=resolved,
            rng=rng,
            metadata_logger=metadata_logger,
        )

    table = workbook_mutations.shift_table_away_from_a1(
        worksheet,
        table,
        top_rows=rng.randint(2, max(2, resolved.max_top_shift)),
        left_columns=rng.randint(1, max(1, resolved.max_left_shift)),
    )
    if resolved.add_title_block:
        title_table = table
        title = resolved.title or f"{resolved.client_name} - PBC schedule"
        table = workbook_mutations.add_title_block(
            worksheet,
            table,
            title=title,
            client_name=resolved.client_name,
            prepared_by=resolved.prepared_by,
            financial_year=resolved.financial_year,
            rng=rng,
            merge_cells=resolved.merge_cells,
            as_at=resolved.report_as_at,
            currency=resolved.report_currency,
        )
        if metadata_logger is not None:
            metadata_logger.record_inserted_note(
                worksheet.title,
                kind="title_block",
                cell=worksheet.cell(title_table.min_row, title_table.min_col).coordinate,
                text=title,
            )
        if resolved.software_signature:
            signature_text = workbook_mutations.add_software_signature(
                worksheet,
                table,
                signature=resolved.software_signature,
                rng=rng,
            )
            if metadata_logger is not None:
                metadata_logger.record_inserted_note(
                    worksheet.title,
                    kind="software_signature",
                    cell=worksheet.cell(max(1, table.min_row - 1), table.min_col).coordinate,
                    text=signature_text,
                )
    if resolved.add_client_notes_block:
        notes_table = table
        table = workbook_mutations.add_client_notes(worksheet, table, rng=rng)
        if metadata_logger is not None:
            metadata_logger.record_inserted_note(
                worksheet.title,
                kind="client_note",
                cell=worksheet.cell(notes_table.min_row, notes_table.min_col).coordinate,
                text=worksheet.cell(notes_table.min_row, notes_table.min_col).value,
            )
    table = workbook_mutations.insert_blank_rows_and_columns(
        worksheet,
        table,
        row_count=resolved.blank_row_count,
        column_count=resolved.blank_column_count,
        rng=rng,
    )
    table = workbook_mutations.duplicate_header_rows(
        worksheet,
        table,
        count=resolved.duplicate_header_count,
        rng=rng,
    )
    if metadata_logger is not None and resolved.duplicate_header_count:
        metadata_logger.record_intentional_error(
            worksheet.title,
            error_type="duplicated_headers",
            details={"count": resolved.duplicate_header_count},
        )
    renamed_columns = workbook_mutations.rename_random_columns(
        worksheet,
        table,
        count=resolved.rename_column_count,
        rng=rng,
        sheet_name=worksheet.title,
    )
    if metadata_logger is not None:
        metadata_logger.record_renamed_columns(worksheet.title, renamed_columns)
    table = workbook_mutations.add_subtotal_rows(
        worksheet,
        table,
        count=resolved.subtotal_row_count,
        rng=rng,
    )
    if metadata_logger is not None and resolved.subtotal_row_count:
        metadata_logger.record_inserted_note(
            worksheet.title,
            kind="subtotal_rows",
            cell=worksheet.cell(table.max_row, table.min_col).coordinate,
            text=f"{resolved.subtotal_row_count} subtotal row(s) inserted",
        )
    table, wrong_period_rows = workbook_mutations.insert_wrong_period_rows(
        worksheet,
        table,
        count=resolved.wrong_period_row_count,
        rng=rng,
    )
    if metadata_logger is not None:
        for row in wrong_period_rows:
            metadata_logger.record_intentional_error(
                worksheet.title,
                error_type="wrong_period_row",
                details=row,
            )
    if resolved.add_footer_notes_block:
        footer_row = table.max_row + 2
        table = workbook_mutations.add_footer_notes(
            worksheet,
            table,
            rng=rng,
            merge_cells=resolved.merge_cells,
        )
        if metadata_logger is not None:
            metadata_logger.record_inserted_note(
                worksheet.title,
                kind="footer_note",
                cell=worksheet.cell(footer_row, table.min_col).coordinate,
                text=worksheet.cell(footer_row, table.min_col).value,
            )

    formatting.apply_inconsistent_formatting(worksheet, table, rng=rng)
    table = formatting.apply_status_cells(worksheet, table, rng=rng)
    formatting.apply_finance_value_formats(worksheet, table, rng=rng)
    stringified_cells = workbook_mutations.stringify_numeric_cells(
        worksheet,
        table,
        count=resolved.stringified_number_count,
        rng=rng,
    )
    if metadata_logger is not None:
        for cell in stringified_cells:
            metadata_logger.record_intentional_error(
                worksheet.title,
                error_type="stringified_number",
                details=cell,
            )
    formula_error_cells = workbook_mutations.inject_formula_errors(
        worksheet,
        table,
        count=resolved.formula_error_count,
        rng=rng,
    )
    if metadata_logger is not None:
        for cell in formula_error_cells:
            metadata_logger.record_intentional_error(
                worksheet.title,
                error_type="formula_error",
                details=cell,
            )
    workbook_mutations.freeze_panes_near_table(worksheet, table)
    workbook_mutations.hide_random_rows_and_columns(
        worksheet,
        table,
        row_count=resolved.hidden_row_count,
        column_count=resolved.hidden_column_count,
        rng=rng,
    )
    comments.add_reviewer_comments(
        worksheet,
        table,
        reviewer_name=resolved.reviewer_name,
        count=resolved.reviewer_comment_count,
        rng=rng,
    )
    secondary_tables = workbook_mutations.add_secondary_tables(
        worksheet,
        table,
        count=resolved.secondary_table_count,
        rng=rng,
    )
    if metadata_logger is not None:
        metadata_logger.set_table_location(worksheet.title, table, worksheet)
        for secondary_table in secondary_tables:
            metadata_logger.record_inserted_note(
                worksheet.title,
                kind="secondary_table",
                cell=secondary_table["start_cell"],
                text=secondary_table,
            )

    if resolved.software_signature:
        bottom = workbook_mutations.find_used_range(worksheet) or table
        footer_anchor = workbook_mutations.TableBounds(
            table.min_row, table.min_col, bottom.max_row, table.max_col
        )
        workbook_mutations.add_report_footer_band(
            worksheet,
            footer_anchor,
            signature=resolved.software_signature,
            prepared_by=resolved.prepared_by,
            reviewer_name=resolved.reviewer_name,
            as_at=resolved.report_as_at,
        )
        if metadata_logger is not None:
            metadata_logger.record_inserted_note(
                worksheet.title,
                kind="report_footer_band",
                cell=worksheet.cell(bottom.max_row + 2, table.min_col).coordinate,
                text=f"System: {resolved.software_signature}",
            )

    if resolved.add_old_version_tabs:
        old_version_tabs = workbook_mutations.add_old_version_tabs(
            workbook,
            source=worksheet,
            count=resolved.old_version_tab_count,
            rng=rng,
        )
        if metadata_logger is not None and old_version_tabs:
            metadata_logger.record_workbook_artifact(
                artifact_type="old_version_tabs",
                details={"source_sheet": worksheet.title, "sheet_names": old_version_tabs},
            )
    if resolved.add_hidden_reconciliation_tabs:
        hidden_sheet = workbook_mutations.add_hidden_reconciliation_tab(workbook, source=worksheet)
        if metadata_logger is not None:
            metadata_logger.record_workbook_artifact(
                artifact_type="hidden_reconciliation_tab",
                details={"source_sheet": worksheet.title, "sheet_name": hidden_sheet.title},
            )

    return workbook
