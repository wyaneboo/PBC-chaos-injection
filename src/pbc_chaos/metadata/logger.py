"""Ground-truth logging for workbook generation."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from hashlib import sha256
from math import isfinite
from secrets import token_hex
from typing import Any

import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet

from pbc_chaos.config_loader import ChaosWorkbookConfig
from pbc_chaos.generators.base import CompanyProfile, FinancialPeriod, GeneratedDocument
from pbc_chaos.metadata.schema import SheetGroundTruth, TableLocation, WorkbookGroundTruth
from pbc_chaos.workbook.workbook_mutations import TableBounds


@dataclass
class _SheetLog:
    document: GeneratedDocument
    sheet_name: str
    clean_records: tuple[dict[str, Any], ...]
    clean_schema: tuple[str, ...]
    original_clean_row_count: int
    table_location: TableLocation | None = None
    renamed_columns_mapping: dict[str, str] = field(default_factory=dict)
    visible_export_profile: str | None = None
    visible_export_department: str | None = None
    visible_export_erp_style: str | None = None
    visible_table_schema: tuple[str, ...] = ()
    visible_column_headers: tuple[str, ...] = ()
    visible_columns_mapping: dict[str, str] = field(default_factory=dict)
    canonical_to_visible_columns: dict[str, str] = field(default_factory=dict)
    omitted_canonical_fields: tuple[str, ...] = ()
    context_field_locations: dict[str, str] = field(default_factory=dict)
    context_field_values: dict[str, Any] = field(default_factory=dict)
    ambiguous_visible_headers: dict[str, str] = field(default_factory=dict)
    inserted_notes: list[dict[str, Any]] = field(default_factory=list)
    intentional_errors: list[dict[str, Any]] = field(default_factory=list)


class GroundTruthLogger:
    """Collects scoring metadata while a workbook is generated and mutated."""

    def __init__(
        self,
        *,
        company: CompanyProfile,
        period: FinancialPeriod,
        config: ChaosWorkbookConfig,
        seed: int | None,
        generated_at: datetime | None = None,
    ) -> None:
        self.company = company
        self.period = period
        self.config = config
        self.seed = seed
        self.generated_at = generated_at or datetime.now(timezone.utc)
        self.workbook_id = _workbook_id(company, period, config, seed)
        self._sheets: dict[str, _SheetLog] = {}
        self._sheet_order: list[str] = []
        self._workbook_intentional_errors: list[dict[str, Any]] = []
        self.injected_discrepancies: list[dict[str, Any]] = []

    def start_sheet(self, document: GeneratedDocument, worksheet: Worksheet) -> None:
        """Register the clean document before layout chaos mutates the sheet."""

        clean_records = tuple(_dataframe_records(document.data))
        clean_schema = tuple(document.metadata.get("expected_canonical_schema", tuple(document.data.columns)))
        self._sheets[worksheet.title] = _SheetLog(
            document=document,
            sheet_name=worksheet.title,
            clean_records=clean_records,
            clean_schema=clean_schema,
            original_clean_row_count=len(document.data),
        )
        self._sheet_order.append(worksheet.title)

    def set_table_location(self, sheet_name: str, table: TableBounds, worksheet: Worksheet) -> None:
        self._sheet(sheet_name).table_location = TableLocation(
            start_cell=worksheet.cell(table.min_row, table.min_col).coordinate,
            end_cell=worksheet.cell(table.max_row, table.max_col).coordinate,
            start_row=table.min_row,
            start_column=table.min_col,
            end_row=table.max_row,
            end_column=table.max_col,
            header_row=table.header_row,
        )

    def record_inserted_note(
        self,
        sheet_name: str,
        *,
        kind: str,
        cell: str,
        text: Any,
    ) -> None:
        self._sheet(sheet_name).inserted_notes.append(
            {"type": kind, "cell": cell, "text": _json_safe(text)}
        )

    def record_renamed_columns(self, sheet_name: str, mapping: dict[str, str]) -> None:
        if mapping:
            self._sheet(sheet_name).renamed_columns_mapping.update(mapping)

    def record_visible_export(self, sheet_name: str, visible_export: Any) -> None:
        """Record workbook-facing export schema metadata for one sheet."""

        sheet = self._sheet(sheet_name)
        sheet.visible_export_profile = str(visible_export.profile_id)
        sheet.visible_export_department = str(visible_export.department)
        sheet.visible_export_erp_style = str(visible_export.erp_style)
        sheet.visible_table_schema = tuple(str(field) for field in visible_export.visible_table_schema)
        sheet.visible_column_headers = tuple(
            str(header) for header in visible_export.visible_column_headers
        )
        sheet.visible_columns_mapping = {
            str(header): str(field)
            for header, field in visible_export.visible_to_canonical.items()
        }
        sheet.canonical_to_visible_columns = {
            str(field): str(header)
            for field, header in visible_export.canonical_to_visible.items()
        }
        sheet.omitted_canonical_fields = tuple(
            str(field) for field in visible_export.omitted_fields
        )
        sheet.context_field_locations = {
            str(field): str(location)
            for field, location in visible_export.context_fields.items()
        }
        sheet.context_field_values = _json_safe(dict(visible_export.context_field_values))
        sheet.ambiguous_visible_headers = {
            str(header): str(field)
            for header, field in visible_export.ambiguous_visible_headers.items()
        }
        self.record_renamed_columns(sheet_name, sheet.canonical_to_visible_columns)

    def record_intentional_error(
        self,
        sheet_name: str,
        *,
        error_type: str,
        details: dict[str, Any],
    ) -> None:
        self._sheet(sheet_name).intentional_errors.append(
            {"type": error_type, **_json_safe(details)}
        )

    def record_workbook_artifact(self, *, artifact_type: str, details: dict[str, Any]) -> None:
        self._workbook_intentional_errors.append({"type": artifact_type, **_json_safe(details)})

    def record_discrepancies(self, discrepancies: list[dict[str, Any]] | tuple[dict[str, Any], ...]) -> None:
        self.injected_discrepancies = [_json_safe(discrepancy) for discrepancy in discrepancies]

    def sheet_names(self) -> tuple[str, ...]:
        """Return primary sheet names tracked in ground-truth metadata."""

        return tuple(self._sheet_order)

    def build_ground_truth(self, workbook) -> WorkbookGroundTruth:
        """Finalize and return the workbook-level ground-truth object."""

        sheet_truth = tuple(
            self._build_sheet_truth(workbook[sheet_name])
            for sheet_name in self._sheet_order
            if sheet_name in workbook.sheetnames
        )
        expected_by_document = {
            sheet.document_type: sheet.expected_extraction_output for sheet in sheet_truth
        }
        document_types = tuple(sheet.document_type for sheet in sheet_truth)
        return WorkbookGroundTruth(
            workbook_id=self.workbook_id,
            company_id=self.company.company_id,
            company_name=self.company.company_name,
            financial_period={
                "financial_year": self.period.financial_year,
                "start_date": self.period.start_date.isoformat(),
                "end_date": self.period.end_date.isoformat(),
            },
            generated_at=self.generated_at.isoformat(),
            seed=self.seed,
            chaos_level={
                "severity": self.config.severity,
                "description": self.config.severity_description,
                "probabilities": self.config.probabilities.as_dict(),
                "unreproducible_nightmare_mode": (
                    self.config.unreproducible_nightmare_mode.as_dict()
                ),
            },
            document_types_included=document_types,
            sheet_names=tuple(workbook.sheetnames),
            clean_canonical_schemas={
                sheet.document_type: sheet.clean_canonical_schema for sheet in sheet_truth
            },
            sheets=sheet_truth,
            injected_discrepancies=tuple(self.injected_discrepancies),
            intentional_errors=tuple(
                _json_safe(error)
                for error in self._workbook_intentional_errors
                + [error for sheet in sheet_truth for error in sheet.intentional_errors]
            ),
            expected_extraction_output=expected_by_document,
        )

    def _build_sheet_truth(self, worksheet: Worksheet) -> SheetGroundTruth:
        log = self._sheet(worksheet.title)
        table = log.table_location or _fallback_table_location(worksheet)
        return SheetGroundTruth(
            sheet_name=worksheet.title,
            document_type=log.document.document_type.value,
            clean_canonical_schema=log.clean_schema,
            original_clean_row_count=log.original_clean_row_count,
            final_messy_row_count=_non_empty_row_count(worksheet),
            table_location=table,
            renamed_columns_mapping=dict(log.renamed_columns_mapping),
            visible_export_profile=log.visible_export_profile,
            visible_export_department=log.visible_export_department,
            visible_export_erp_style=log.visible_export_erp_style,
            visible_table_schema=log.visible_table_schema,
            visible_column_headers=log.visible_column_headers,
            visible_columns_mapping=dict(log.visible_columns_mapping),
            canonical_to_visible_columns=dict(log.canonical_to_visible_columns),
            omitted_canonical_fields=log.omitted_canonical_fields,
            context_field_locations=dict(log.context_field_locations),
            context_field_values=dict(log.context_field_values),
            ambiguous_visible_headers=dict(log.ambiguous_visible_headers),
            hidden_rows=tuple(
                row_idx for row_idx, dimension in worksheet.row_dimensions.items() if dimension.hidden
            ),
            hidden_columns=tuple(
                col_letter
                for col_letter, dimension in worksheet.column_dimensions.items()
                if dimension.hidden
            ),
            merged_cell_ranges=tuple(str(cell_range) for cell_range in worksheet.merged_cells.ranges),
            inserted_notes=tuple(log.inserted_notes),
            intentional_errors=tuple(_json_safe(error) for error in log.intentional_errors),
            expected_extraction_output=log.clean_records,
        )

    def _sheet(self, sheet_name: str) -> _SheetLog:
        try:
            return self._sheets[sheet_name]
        except KeyError as exc:
            raise KeyError(f"No ground-truth sheet log exists for {sheet_name!r}.") from exc


def json_safe(value: Any) -> Any:
    """Return a JSON-compatible version of nested Python/pandas/openpyxl values."""

    return _json_safe(value)


def _workbook_id(
    company: CompanyProfile,
    period: FinancialPeriod,
    config: ChaosWorkbookConfig,
    seed: int | None,
) -> str:
    payload_parts = [company.company_id, str(period.financial_year), str(config.severity), str(seed)]
    if config.unreproducible_nightmare_mode.enabled:
        payload_parts.extend(("unreproducible", token_hex(8)))
    payload = ":".join(payload_parts).encode("utf-8")
    return f"wb_{sha256(payload).hexdigest()[:16]}"


def _dataframe_records(data: pd.DataFrame) -> list[dict[str, Any]]:
    records = data.to_dict(orient="records")
    return [_json_safe(record) for record in records]


def _json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, list | tuple):
        return [_json_safe(item) for item in value]
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is pd.NA:
        return None
    if isinstance(value, float) and not isfinite(value):
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if hasattr(value, "item"):
        try:
            return _json_safe(value.item())
        except (TypeError, ValueError):
            pass
    return value


def _fallback_table_location(worksheet: Worksheet) -> TableLocation:
    min_row = worksheet.min_row
    min_col = worksheet.min_column
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    return TableLocation(
        start_cell=worksheet.cell(min_row, min_col).coordinate,
        end_cell=worksheet.cell(max_row, max_col).coordinate,
        start_row=min_row,
        start_column=min_col,
        end_row=max_row,
        end_column=max_col,
        header_row=min_row,
    )


def _non_empty_row_count(worksheet: Worksheet) -> int:
    count = 0
    for row in worksheet.iter_rows():
        if any(cell.value is not None for cell in row):
            count += 1
    return count
