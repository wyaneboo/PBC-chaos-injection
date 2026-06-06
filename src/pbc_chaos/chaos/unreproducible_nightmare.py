"""Non-deterministic AI-style post-pass for nightmare workbooks."""

from __future__ import annotations

from dataclasses import dataclass
import json
import os
from random import SystemRandom
from typing import Any, Callable, TYPE_CHECKING, TypedDict
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen
import warnings

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from pbc_chaos.config_loader import UnreproducibleNightmareModeConfig
from pbc_chaos.env import load_env_file
from pbc_chaos.workbook import pbc_tracker_layout
from pbc_chaos.workbook import workbook_mutations

if TYPE_CHECKING:
    from pbc_chaos.metadata.logger import GroundTruthLogger


class NightmareAgentState(TypedDict, total=False):
    workbook: Workbook
    allowed_sheet_names: tuple[str, ...]
    summary: dict[str, Any]
    plan: list[dict[str, Any]]
    planner_provider: str
    planner_error: str | None


@dataclass(frozen=True)
class NightmareAgentResult:
    provider: str
    plan: tuple[dict[str, Any], ...]
    error: str | None = None


NightmareProgressCallback = Callable[[dict[str, Any]], None]


class UnreproducibleNightmareAgent:
    """LangGraph agent that reviews a messy workbook and chooses extra realism actions."""

    def __init__(self, config: UnreproducibleNightmareModeConfig) -> None:
        self.config = config
        self.rng = SystemRandom()

    def plan(
        self,
        workbook: Workbook,
        *,
        allowed_sheet_names: tuple[str, ...] = (),
    ) -> NightmareAgentResult:
        initial_state: NightmareAgentState = {
            "workbook": workbook,
            "allowed_sheet_names": allowed_sheet_names,
        }
        graph = self._build_langgraph()
        planned_state = graph.invoke(initial_state)
        return NightmareAgentResult(
            provider=_langgraph_provider(planned_state.get("planner_provider")),
            plan=tuple(planned_state.get("plan", ())),
            error=planned_state.get("planner_error"),
        )

    def _build_langgraph(self) -> Any:
        with warnings.catch_warnings():
            from langchain_core._api.deprecation import LangChainPendingDeprecationWarning

            warnings.filterwarnings("ignore", category=LangChainPendingDeprecationWarning)
            from langgraph.graph import END, START, StateGraph

        builder = StateGraph(NightmareAgentState)
        builder.add_node("review_workbook", self._review_node)
        builder.add_node("choose_chaos_tools", self._plan_node)
        builder.add_edge(START, "review_workbook")
        builder.add_edge("review_workbook", "choose_chaos_tools")
        builder.add_edge("choose_chaos_tools", END)
        return builder.compile()

    def _review_node(self, state: NightmareAgentState) -> NightmareAgentState:
        workbook = state["workbook"]
        allowed_sheet_names = set(state.get("allowed_sheet_names") or ())
        sheet_summaries = []
        for worksheet in workbook.worksheets:
            if allowed_sheet_names and worksheet.title not in allowed_sheet_names:
                continue
            table = workbook_mutations.find_used_range(worksheet)
            sheet_summaries.append(
                {
                    "sheet_name": worksheet.title,
                    "hidden": worksheet.sheet_state != "visible",
                    "used_rows": 0 if table is None else table.max_row - table.min_row + 1,
                    "used_columns": 0 if table is None else table.max_col - table.min_col + 1,
                    "comments": _comment_count(worksheet),
                    "formulas": _formula_count(worksheet),
                    "merged_ranges": len(tuple(worksheet.merged_cells.ranges)),
                    "hidden_rows": sum(
                        1 for dimension in worksheet.row_dimensions.values() if dimension.hidden
                    ),
                    "hidden_columns": sum(
                        1 for dimension in worksheet.column_dimensions.values() if dimension.hidden
                    ),
                }
            )
        return {
            "summary": {
                "sheet_count": len(workbook.worksheets),
                "visible_sheet_names": [
                    item["sheet_name"] for item in sheet_summaries if not item["hidden"]
                ],
                "hidden_sheet_count": sum(1 for item in sheet_summaries if item["hidden"]),
                "sheets": sheet_summaries,
            }
        }

    def _plan_node(self, state: NightmareAgentState) -> NightmareAgentState:
        summary = state["summary"]
        visible_sheet_names = tuple(summary.get("visible_sheet_names") or ())
        if not visible_sheet_names:
            return {"plan": [], "planner_provider": "no_visible_sheets", "planner_error": None}

        llm_plan, llm_error = self._llm_plan(summary)
        if llm_plan:
            return {
                "plan": llm_plan,
                "planner_provider": "gemini_generate_content",
                "planner_error": None,
            }

        plan: list[dict[str, Any]] = []
        notation_allocations: dict[str, int] = {}
        for _ in range(self.config.notation_count):
            sheet_name = self.rng.choice(visible_sheet_names)
            notation_allocations[sheet_name] = notation_allocations.get(sheet_name, 0) + 1
        for sheet_name, count in sorted(notation_allocations.items()):
            plan.append({"tool": "human_residue_notation", "sheet": sheet_name, "count": count})

        remaining_extra_tools = self.config.extra_tool_count
        if (
            remaining_extra_tools > 0
            and pbc_tracker_layout.TRACKER_SHEET_NAME in visible_sheet_names
        ):
            plan.append(
                {
                    "tool": "add_visible_tracker_comment",
                    "sheet": pbc_tracker_layout.TRACKER_SHEET_NAME,
                    "text": self.rng.choice(
                        (
                            "Not provided, remind finance",
                            "Hardcopy only so far",
                            "Need follow-up before audit review",
                        )
                    ),
                }
            )
            remaining_extra_tools -= 1

        for _ in range(remaining_extra_tools):
            tool = self._choose_extra_tool(summary)
            action: dict[str, Any] = {"tool": tool}
            if tool in _TRACKER_TOOLS:
                action["sheet"] = pbc_tracker_layout.TRACKER_SHEET_NAME
            elif tool not in {"hidden_reconciliation_tab"}:
                action["sheet"] = self.rng.choice(visible_sheet_names)
            if tool in {"formula_errors", "stringified_numbers"}:
                action["count"] = self.rng.randint(1, 3)
            elif tool == "hide_rows_columns":
                action["row_count"] = self.rng.randint(1, 3)
                action["column_count"] = self.rng.randint(0, 1)
            else:
                action["count"] = 1
            plan.append(action)

        return {
            "plan": plan,
            "planner_provider": "heuristic_fallback",
            "planner_error": llm_error,
        }

    def _llm_plan(self, summary: dict[str, Any]) -> tuple[list[dict[str, Any]], str | None]:
        if not self.config.use_llm_planner:
            return [], "LLM planner disabled by config."

        load_env_file()
        api_key = os.getenv(self.config.gemini_api_key_env)
        if not api_key:
            return [], f"{self.config.gemini_api_key_env} is not set."

        prompt = _planner_prompt(
            summary=summary,
            notation_count=self.config.notation_count,
            extra_tool_count=self.config.extra_tool_count,
        )
        payload = {
            "systemInstruction": {
                "parts": [
                    {
                        "text": (
                            "You are a spreadsheet chaos reviewer for synthetic audit PBC "
                            "workbooks. Return only a valid JSON object matching the schema."
                        )
                    }
                ]
            },
            "contents": [
                {"role": "user", "parts": [{"text": prompt}]},
            ],
            "generationConfig": {
                "responseMimeType": "application/json",
                "responseJsonSchema": _plan_response_schema(),
                "maxOutputTokens": 1800,
            },
        }

        try:
            response = _post_gemini_generate_content(
                model=self.config.llm_model,
                payload=payload,
                api_key=api_key,
                timeout=self.config.llm_timeout_seconds,
            )
            parsed = json.loads(_response_text(response))
        except (OSError, HTTPError, URLError, TimeoutError, json.JSONDecodeError, KeyError) as exc:
            return [], f"LLM planner failed: {exc}"

        raw_plan = parsed.get("plan") if isinstance(parsed, dict) else None
        if not isinstance(raw_plan, list):
            return [], "LLM planner did not return a plan list."
        return _sanitize_plan(
            raw_plan,
            summary=summary,
            notation_count=self.config.notation_count,
            extra_tool_count=self.config.extra_tool_count,
        ), None

    def _choose_extra_tool(self, summary: dict[str, Any]) -> str:
        hidden_sheet_count = int(summary.get("hidden_sheet_count") or 0)
        tools = [
            "formula_errors",
            "stringified_numbers",
            "hide_rows_columns",
            "secondary_table",
            "old_version_tab",
        ]
        if hidden_sheet_count == 0:
            tools.extend(["hidden_reconciliation_tab", "hidden_reconciliation_tab"])
        if pbc_tracker_layout.TRACKER_SHEET_NAME in summary.get("visible_sheet_names", ()):
            tools.extend(
                [
                    "add_visible_tracker_comment",
                    "apply_tracker_status_variant",
                    "apply_tracker_deadline_noise",
                    "highlight_tracker_update_row",
                    "apply_tracker_follow_up_noise",
                ]
            )
        return self.rng.choice(tools)


_SHEET_TOOLS = {
    "human_residue_notation",
    "formula_errors",
    "stringified_numbers",
    "hide_rows_columns",
    "secondary_table",
    "old_version_tab",
    "add_visible_tracker_comment",
    "apply_tracker_status_variant",
    "apply_tracker_deadline_noise",
    "highlight_tracker_update_row",
    "apply_tracker_follow_up_noise",
}
_WORKBOOK_TOOLS = {"hidden_reconciliation_tab"}
_ALL_TOOLS = _SHEET_TOOLS | _WORKBOOK_TOOLS
_TRACKER_TOOLS = {
    "add_visible_tracker_comment",
    "apply_tracker_status_variant",
    "apply_tracker_deadline_noise",
    "highlight_tracker_update_row",
    "apply_tracker_follow_up_noise",
}


def _planner_prompt(
    *,
    summary: dict[str, Any],
    notation_count: int,
    extra_tool_count: int,
) -> str:
    return json.dumps(
        {
            "task": (
                "Review this already messy generated PBC workbook summary. Choose a realistic "
                "post-pass chaos plan. Human residue notations are the priority: they simulate "
                "finance staff using spreadsheet cells/comments to remind themselves or pass "
                "information to another person."
            ),
            "constraints": {
                "allowed_tools": sorted(_ALL_TOOLS),
                "sheet_specific_tools": sorted(_SHEET_TOOLS),
                "workbook_tools": sorted(_WORKBOOK_TOOLS),
                "required_human_residue_notation_count": notation_count,
                "target_extra_tool_count": extra_tool_count,
                "available_sheet_names": summary.get("visible_sheet_names", []),
                    "notes": [
                        "Use only available sheet names.",
                        "Use tracker tools only on the PBC Request List sheet.",
                        "For tracker tools, row_key can be a visible request id such as A.7.",
                        "Prefer believable audit-season mess, not file corruption.",
                        "Do not request direct accounting truth changes.",
                        "Keep counts small and plausible.",
                ],
            },
            "workbook_summary": summary,
            "response_shape": {
                "plan": [
                    {
                        "tool": "human_residue_notation",
                        "sheet": "Trial Balance",
                        "count": 3,
                        "texts": [
                            "pls confirm AP cutoff before sending",
                            "old support in email, do not delete",
                        ],
                        "reason": "Staff reminder notes are sparse on this sheet.",
                    },
                    {
                        "tool": "add_visible_tracker_comment",
                        "sheet": "PBC Request List",
                        "row_key": "A.7",
                        "text": "Not provided, remind finance",
                        "reason": "Tracker row needs a visible follow-up note.",
                    }
                ]
            },
        },
        indent=2,
        sort_keys=True,
    )


def _plan_response_schema() -> dict[str, Any]:
    return {
        "type": "object",
        "properties": {
            "plan": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "tool": {"type": "string", "enum": sorted(_ALL_TOOLS)},
                        "sheet": {"type": "string"},
                        "count": {"type": "integer", "minimum": 0, "maximum": 20},
                        "row_count": {"type": "integer", "minimum": 0, "maximum": 5},
                        "column_count": {"type": "integer", "minimum": 0, "maximum": 3},
                        "texts": {
                            "type": "array",
                            "items": {"type": "string"},
                            "maxItems": 20,
                        },
                        "row_key": {"type": "string"},
                        "text": {"type": "string"},
                        "status": {"type": "string"},
                        "style": {"type": "string"},
                        "reason": {"type": "string"},
                    },
                    "required": ["tool"],
                    "additionalProperties": False,
                },
            }
        },
        "required": ["plan"],
        "additionalProperties": False,
    }


def _post_gemini_generate_content(
    *,
    model: str,
    payload: dict[str, Any],
    api_key: str,
    timeout: int,
) -> dict[str, Any]:
    data = json.dumps(payload).encode("utf-8")
    request = Request(
        f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent",
        data=data,
        headers={
            "x-goog-api-key": api_key,
            "Content-Type": "application/json",
        },
        method="POST",
    )
    with urlopen(request, timeout=timeout) as response:
        return json.loads(response.read().decode("utf-8"))


def _response_text(response: dict[str, Any]) -> str:
    fragments: list[str] = []
    for candidate in response.get("candidates", []):
        if not isinstance(candidate, dict):
            continue
        content = candidate.get("content")
        if not isinstance(content, dict):
            continue
        for part in content.get("parts", []):
            if isinstance(part, dict) and isinstance(part.get("text"), str):
                fragments.append(part["text"])
    if not fragments:
        raise KeyError("No text returned by Gemini generateContent API.")
    return "".join(fragments)


def _sanitize_plan(
    raw_plan: list[Any],
    *,
    summary: dict[str, Any],
    notation_count: int,
    extra_tool_count: int,
) -> list[dict[str, Any]]:
    visible_sheet_names = tuple(str(name) for name in summary.get("visible_sheet_names", ()) if name)
    if not visible_sheet_names:
        return []

    sanitized: list[dict[str, Any]] = []
    notation_total = 0
    extra_total = 0
    sheet_index = 0
    for raw_action in raw_plan[: max(10, notation_count + extra_tool_count + 5)]:
        if not isinstance(raw_action, dict):
            continue
        tool = str(raw_action.get("tool", "")).strip()
        if tool not in _ALL_TOOLS:
            continue

        action: dict[str, Any] = {"tool": tool}
        if tool in _SHEET_TOOLS:
            sheet_name = str(raw_action.get("sheet", "")).strip()
            if tool in _TRACKER_TOOLS:
                if pbc_tracker_layout.TRACKER_SHEET_NAME not in visible_sheet_names:
                    continue
                sheet_name = pbc_tracker_layout.TRACKER_SHEET_NAME
            elif sheet_name not in visible_sheet_names:
                sheet_name = visible_sheet_names[sheet_index % len(visible_sheet_names)]
                sheet_index += 1
            action["sheet"] = sheet_name

        if tool == "human_residue_notation":
            count = _bounded_int(raw_action.get("count"), default=1, minimum=1, maximum=10)
            notation_total += count
            action["count"] = count
            texts = raw_action.get("texts")
            if isinstance(texts, list):
                cleaned_texts = [
                    text.strip()[:160]
                    for text in texts
                    if isinstance(text, str) and text.strip()
                ]
                if cleaned_texts:
                    action["texts"] = tuple(cleaned_texts[:count])
        elif tool in _TRACKER_TOOLS:
            if extra_total >= extra_tool_count:
                continue
            for key, max_length in (
                ("row_key", 24),
                ("text", 120),
                ("status", 48),
                ("style", 48),
            ):
                value = raw_action.get(key)
                if isinstance(value, str) and value.strip():
                    action[key] = value.strip()[:max_length]
            extra_total += 1
        elif tool == "hide_rows_columns":
            if extra_total >= extra_tool_count:
                continue
            action["row_count"] = _bounded_int(raw_action.get("row_count"), default=1, minimum=0, maximum=3)
            action["column_count"] = _bounded_int(
                raw_action.get("column_count"),
                default=0,
                minimum=0,
                maximum=1,
            )
            extra_total += 1
        else:
            if extra_total >= extra_tool_count:
                continue
            action["count"] = _bounded_int(raw_action.get("count"), default=1, minimum=1, maximum=3)
            extra_total += 1

        reason = raw_action.get("reason")
        if isinstance(reason, str) and reason.strip():
            action["reason"] = reason.strip()[:160]
        sanitized.append(action)

    while notation_total < notation_count:
        sheet_name = visible_sheet_names[sheet_index % len(visible_sheet_names)]
        sheet_index += 1
        count = min(10, notation_count - notation_total)
        sanitized.append({"tool": "human_residue_notation", "sheet": sheet_name, "count": count})
        notation_total += count

    filler_tools = ["formula_errors", "stringified_numbers", "hide_rows_columns", "secondary_table"]
    if pbc_tracker_layout.TRACKER_SHEET_NAME in visible_sheet_names:
        filler_tools.extend(["add_visible_tracker_comment", "apply_tracker_follow_up_noise"])
    while extra_total < extra_tool_count:
        tool = filler_tools[extra_total % len(filler_tools)]
        if tool in _TRACKER_TOOLS:
            sheet_name = pbc_tracker_layout.TRACKER_SHEET_NAME
        else:
            sheet_name = visible_sheet_names[sheet_index % len(visible_sheet_names)]
            sheet_index += 1
        action = {"tool": tool, "sheet": sheet_name}
        if tool == "hide_rows_columns":
            action["row_count"] = 1
            action["column_count"] = 0
        elif tool not in _TRACKER_TOOLS:
            action["count"] = 1
        sanitized.append(action)
        extra_total += 1

    return sanitized


def _bounded_int(value: Any, *, default: int, minimum: int, maximum: int) -> int:
    if isinstance(value, bool):
        return default
    if not isinstance(value, int):
        return default
    return max(minimum, min(maximum, value))


def _langgraph_provider(provider: str | None) -> str:
    if not provider:
        return "langgraph_heuristic"
    if provider.startswith("langgraph_"):
        return provider
    return f"langgraph_{provider}"


def apply_unreproducible_nightmare_mode(
    *,
    workbook: Workbook,
    logger: GroundTruthLogger,
    config: UnreproducibleNightmareModeConfig,
    progress_callback: NightmareProgressCallback | None = None,
) -> NightmareAgentResult:
    """Run the optional non-deterministic nightmare post-pass."""

    agent = UnreproducibleNightmareAgent(config)
    _notify_progress(
        progress_callback,
        phase="reviewing",
        message="Reviewing workbook structure and sheet metadata",
        agent_provider="pending",
        planned_actions=(),
        applied_actions=(),
    )
    result = agent.plan(workbook, allowed_sheet_names=logger.sheet_names())
    rng = agent.rng
    allowed_sheet_names = logger.sheet_names()

    applied_actions = []
    _notify_progress(
        progress_callback,
        phase="planned",
        message=f"Selected {len(result.plan)} spreadsheet change(s)",
        agent_provider=result.provider,
        agent_error=result.error,
        planned_actions=result.plan,
        applied_actions=tuple(applied_actions),
    )
    for action_index, action in enumerate(result.plan, start=1):
        applied = _apply_action(
            workbook=workbook,
            logger=logger,
            action=action,
            config=config,
            rng=rng,
            allowed_sheet_names=allowed_sheet_names,
        )
        if applied:
            applied_actions.append(applied)
            _notify_progress(
                progress_callback,
                phase="applied",
                message=_action_message(applied, action_index=action_index, action_total=len(result.plan)),
                agent_provider=result.provider,
                agent_error=result.error,
                planned_actions=result.plan,
                applied_actions=tuple(applied_actions),
                current_action=applied,
                action_index=action_index,
                action_total=len(result.plan),
            )

    logger.record_workbook_artifact(
        artifact_type="unreproducible_nightmare_plan",
        details={
            "agent_provider": result.provider,
            "agent_error": result.error,
            "planned_actions": result.plan,
            "applied_actions": tuple(applied_actions),
        },
    )
    _notify_progress(
        progress_callback,
        phase="complete",
        message=f"Completed {len(applied_actions)} spreadsheet change(s)",
        agent_provider=result.provider,
        agent_error=result.error,
        planned_actions=result.plan,
        applied_actions=tuple(applied_actions),
    )
    return result


def _notify_progress(
    callback: NightmareProgressCallback | None,
    *,
    phase: str,
    message: str,
    agent_provider: str,
    planned_actions: tuple[dict[str, Any], ...],
    applied_actions: tuple[dict[str, Any], ...],
    agent_error: str | None = None,
    current_action: dict[str, Any] | None = None,
    action_index: int | None = None,
    action_total: int | None = None,
) -> None:
    if callback is None:
        return
    callback(
        {
            "phase": phase,
            "message": message,
            "agent_provider": agent_provider,
            "agent_error": agent_error,
            "planned_actions": planned_actions,
            "applied_actions": applied_actions,
            "planned_action_count": len(planned_actions),
            "applied_action_count": len(applied_actions),
            "current_action": current_action,
            "action_index": action_index,
            "action_total": action_total,
        }
    )


def _action_message(action: dict[str, Any], *, action_index: int, action_total: int) -> str:
    tool = str(action.get("tool") or "agent action").replace("_", " ")
    sheet = action.get("sheet") or action.get("source_sheet") or action.get("sheet_name")
    target = f" on {sheet}" if sheet else ""
    return f"Applied {action_index}/{action_total}: {tool}{target}"


def _apply_action(
    *,
    workbook: Workbook,
    logger: GroundTruthLogger,
    action: dict[str, Any],
    config: UnreproducibleNightmareModeConfig,
    rng: SystemRandom,
    allowed_sheet_names: tuple[str, ...],
) -> dict[str, Any] | None:
    tool = str(action.get("tool", ""))
    if tool == "hidden_reconciliation_tab":
        source = _random_visible_sheet(workbook, rng, allowed_sheet_names=allowed_sheet_names)
        if source is None:
            return None
        hidden_sheet = workbook_mutations.add_hidden_reconciliation_tab(workbook, source=source)
        logger.record_workbook_artifact(
            artifact_type="ai_hidden_reconciliation_tab",
            details={"source_sheet": source.title, "sheet_name": hidden_sheet.title},
        )
        return {"tool": tool, "source_sheet": source.title, "sheet_name": hidden_sheet.title}

    sheet_name = str(action.get("sheet", ""))
    if sheet_name not in allowed_sheet_names or sheet_name not in workbook.sheetnames:
        return None
    worksheet = workbook[sheet_name]

    if tool in _TRACKER_TOOLS:
        return pbc_tracker_layout.apply_tracker_agent_action(
            worksheet=worksheet,
            action=action,
            logger=logger,
            rng=rng,
        )

    table = workbook_mutations.find_used_range(worksheet)
    if table is None:
        return None

    if tool == "human_residue_notation":
        placed = _add_random_notations(
            worksheet,
            table=table,
            logger=logger,
            count=int(action.get("count") or 0),
            max_length=config.max_notation_length,
            text_pool=tuple(
                text
                for text in action.get("texts", ())
                if isinstance(text, str) and text.strip()
            ),
            rng=rng,
        )
        return {"tool": tool, "sheet": sheet_name, "count": len(placed), "cells": tuple(placed)}

    if tool == "formula_errors":
        changed = workbook_mutations.inject_formula_errors(
            worksheet,
            table,
            count=int(action.get("count") or 0),
            rng=rng,
        )
        for cell in changed:
            logger.record_intentional_error(sheet_name, error_type="ai_formula_error", details=cell)
        return {"tool": tool, "sheet": sheet_name, "count": len(changed)}

    if tool == "stringified_numbers":
        changed = workbook_mutations.stringify_numeric_cells(
            worksheet,
            table,
            count=int(action.get("count") or 0),
            rng=rng,
        )
        for cell in changed:
            logger.record_intentional_error(
                sheet_name,
                error_type="ai_stringified_number",
                details=cell,
            )
        return {"tool": tool, "sheet": sheet_name, "count": len(changed)}

    if tool == "hide_rows_columns":
        row_count = int(action.get("row_count") or 0)
        column_count = int(action.get("column_count") or 0)
        workbook_mutations.hide_random_rows_and_columns(
            worksheet,
            table,
            row_count=row_count,
            column_count=column_count,
            rng=rng,
        )
        logger.record_intentional_error(
            sheet_name,
            error_type="ai_hidden_rows_columns",
            details={"row_count": row_count, "column_count": column_count},
        )
        return {
            "tool": tool,
            "sheet": sheet_name,
            "row_count": row_count,
            "column_count": column_count,
        }

    if tool == "secondary_table":
        secondary_tables = workbook_mutations.add_secondary_tables(
            worksheet,
            table,
            count=int(action.get("count") or 0),
            rng=rng,
        )
        for secondary_table in secondary_tables:
            logger.record_inserted_note(
                sheet_name,
                kind="ai_secondary_table",
                cell=secondary_table["start_cell"],
                text=secondary_table,
            )
        return {"tool": tool, "sheet": sheet_name, "count": len(secondary_tables)}

    if tool == "old_version_tab":
        old_tabs = workbook_mutations.add_old_version_tabs(
            workbook,
            source=worksheet,
            count=int(action.get("count") or 0),
            rng=rng,
        )
        if old_tabs:
            logger.record_workbook_artifact(
                artifact_type="ai_old_version_tabs",
                details={"source_sheet": sheet_name, "sheet_names": old_tabs},
            )
        return {"tool": tool, "sheet": sheet_name, "sheet_names": tuple(old_tabs)}

    return None


def _add_random_notations(
    worksheet: Worksheet,
    *,
    table: workbook_mutations.TableBounds,
    logger: GroundTruthLogger,
    count: int,
    max_length: int,
    text_pool: tuple[str, ...] = (),
    rng: SystemRandom,
) -> list[str]:
    placed: list[str] = []
    if count <= 0:
        return placed

    for _ in range(count):
        mode = rng.choice(("comment", "visible_note", "visible_note"))
        text = (rng.choice(text_pool) if text_pool else _notation_text(rng))[:max_length]
        author = rng.choice(("Finance", "AP team", "Audit", "CY", "Jason", "Mei", "Client"))
        if mode == "comment":
            cell = _random_non_empty_cell(worksheet, table, rng)
            if cell is None:
                continue
            if cell.comment is None:
                cell.comment = Comment(text, author)
            else:
                cell.comment = Comment(f"{cell.comment.text}\n{text}", cell.comment.author or author)
        else:
            cell = _random_blank_cell(worksheet, table, rng)
            if cell is None:
                continue
            cell.value = text
            cell.fill = PatternFill("solid", fgColor=rng.choice(("FFF2CC", "FCE4D6", "E2F0D9")))
            cell.font = Font(italic=True, color="666666")
            cell.alignment = Alignment(wrap_text=True)

        logger.record_inserted_note(
            worksheet.title,
            kind="ai_random_notation",
            cell=cell.coordinate,
            text={"mode": mode, "author": author, "text": text},
        )
        placed.append(cell.coordinate)

    return placed


def _random_non_empty_cell(
    worksheet: Worksheet,
    table: workbook_mutations.TableBounds,
    rng: SystemRandom,
):
    candidates = [
        cell
        for row in worksheet.iter_rows(
            min_row=table.min_row,
            max_row=table.max_row,
            min_col=table.min_col,
            max_col=table.max_col,
        )
        for cell in row
        if cell.value is not None and not isinstance(cell, MergedCell)
    ]
    if not candidates:
        return None
    return rng.choice(candidates)


def _random_blank_cell(
    worksheet: Worksheet,
    table: workbook_mutations.TableBounds,
    rng: SystemRandom,
):
    max_row = table.max_row + 8
    max_col = table.max_col + 4
    candidates = []
    for _ in range(200):
        row = rng.randint(1, max_row)
        col = rng.randint(1, max_col)
        cell = worksheet.cell(row, col)
        if cell.value is None and not isinstance(cell, MergedCell):
            candidates.append(cell)
    if not candidates:
        return None
    return rng.choice(candidates)


def _notation_text(rng: SystemRandom) -> str:
    prefixes = (
        "check",
        "pls confirm",
        "pending",
        "use latest",
        "not sure",
        "revised",
        "ask client",
        "hold first",
    )
    subjects = (
        "GL tie-out",
        "AP accrual",
        "bank recon",
        "audit sample",
        "old support",
        "manual JV",
        "cutoff item",
        "tax code",
        "variance",
        "FY close",
    )
    suffixes = (
        "before sending",
        "after manager review",
        "per Teams chat",
        "not tally to TB",
        "support in email",
        "waiting for final file",
        "maybe old version",
        "do not delete",
        "confirm with finance",
        "update later",
    )
    return f"{rng.choice(prefixes)} {rng.choice(subjects)} - {rng.choice(suffixes)}"


def _random_visible_sheet(
    workbook: Workbook,
    rng: SystemRandom,
    *,
    allowed_sheet_names: tuple[str, ...],
) -> Worksheet | None:
    allowed = set(allowed_sheet_names)
    visible = [
        sheet
        for sheet in workbook.worksheets
        if sheet.sheet_state == "visible" and sheet.title in allowed
    ]
    if not visible:
        return None
    return rng.choice(visible)


def _comment_count(worksheet: Worksheet) -> int:
    return sum(1 for row in worksheet.iter_rows() for cell in row if cell.comment is not None)


def _formula_count(worksheet: Worksheet) -> int:
    return sum(
        1
        for row in worksheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )
