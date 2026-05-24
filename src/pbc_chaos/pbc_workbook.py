"""High-level workbook generation API."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Mapping

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from pbc_chaos.config_loader import ChaosWorkbookConfig, config_from_mapping, load_config
from pbc_chaos.generators.base import CompanyProfile, FinancialPeriod, GeneratedDocument
from pbc_chaos.generators.expense_claims import ExpenseClaimListingGenerator
from pbc_chaos.metadata.logger import GroundTruthLogger
from pbc_chaos.metadata.schema import WorkbookGroundTruth
from pbc_chaos.reconciliation import (
    ReconciliationContext,
    generate_ap_aging,
    generate_ar_aging,
    generate_bank_reconciliation,
    generate_fixed_asset_register,
    generate_general_ledger,
    generate_inventory_listing,
    generate_journal_entry_listing,
    generate_payroll_detail,
    generate_payroll_summary,
    generate_trial_balance,
)
from pbc_chaos.workbook import layout_engine


SHEET_NAMES = {
    "trial_balance": "Trial Balance",
    "general_ledger": "General Ledger",
    "ap_aging": "AP Aging",
    "ar_aging": "AR Aging",
    "bank_reconciliation": "Bank Recon",
    "fixed_asset_register": "Fixed Assets",
    "payroll_summary": "Payroll Summary",
    "payroll_detail": "Payroll Detail",
    "inventory_listing": "Inventory",
    "journal_entry_listing": "Journal Entries",
    "expense_claim_listing": "Expense Claims",
}


@dataclass(frozen=True)
class GeneratedPBCWorkbook:
    workbook: Workbook
    ground_truth: WorkbookGroundTruth


def generate_pbc_workbook(
    company: CompanyProfile,
    period: FinancialPeriod,
    *,
    config: ChaosWorkbookConfig | Mapping[str, object] | str | Path | None = None,
    seed: int | None = None,
) -> Workbook:
    """Generate one workbook using the configured Phase 7 chaos severity."""

    return generate_pbc_workbook_with_ground_truth(
        company,
        period,
        config=config,
        seed=seed,
    ).workbook


def generate_pbc_workbook_with_ground_truth(
    company: CompanyProfile,
    period: FinancialPeriod,
    *,
    config: ChaosWorkbookConfig | Mapping[str, object] | str | Path | None = None,
    seed: int | None = None,
) -> GeneratedPBCWorkbook:
    """Generate a workbook and matching machine-readable ground truth."""

    resolved = _coerce_config(config)
    context, documents = _generate_documents(company, period, seed=seed)
    logger = GroundTruthLogger(company=company, period=period, config=resolved, seed=seed)

    workbook = Workbook()
    workbook.remove(workbook.active)
    for index, document in enumerate(documents, start=1):
        sheet_name = SHEET_NAMES[document.document_type.value]
        worksheet = workbook.create_sheet(sheet_name)
        for row in dataframe_to_rows(document.data, index=False, header=True):
            worksheet.append(row)
        logger.start_sheet(document, worksheet)

        layout_config = resolved.layout_config(
            company=company,
            period=period,
            title=f"{sheet_name} - PBC support",
            seed=seed,
            sheet_index=index,
            hidden_recon_allowed=True,
        )
        layout_engine.apply_layout_chaos(
            workbook=workbook,
            worksheet=worksheet,
            config=layout_config,
            seed=(0 if seed is None else seed) + index,
            metadata_logger=logger,
        )

    workbook.properties.title = f"{company.company_name} PBC workbook"
    workbook.properties.subject = (
        f"Chaos severity {resolved.severity}: {resolved.severity_description}"
    )
    logger.record_discrepancies(context.discrepancy_metadata)
    return GeneratedPBCWorkbook(
        workbook=workbook,
        ground_truth=logger.build_ground_truth(workbook),
    )


def _coerce_config(
    config: ChaosWorkbookConfig | Mapping[str, object] | str | Path | None,
) -> ChaosWorkbookConfig:
    if config is None:
        return config_from_mapping({})
    if isinstance(config, ChaosWorkbookConfig):
        return config
    if isinstance(config, str | Path):
        return load_config(config)
    return config_from_mapping(config)


def _generate_documents(
    company: CompanyProfile,
    period: FinancialPeriod,
    *,
    seed: int | None,
) -> tuple[ReconciliationContext, tuple[GeneratedDocument, ...]]:
    context = ReconciliationContext(company, period, seed=seed)
    documents = [
        generate_trial_balance(context),
        generate_general_ledger(context),
        generate_bank_reconciliation(context),
        generate_ap_aging(context),
        generate_ar_aging(context),
        generate_fixed_asset_register(context),
        generate_payroll_summary(context),
        generate_payroll_detail(context),
        generate_inventory_listing(context),
        generate_journal_entry_listing(context),
    ]
    documents.append(
        ExpenseClaimListingGenerator().generate(
            company=company,
            period=period,
            seed=(0 if seed is None else seed) + 50_000,
        )
    )
    return context, tuple(documents)
